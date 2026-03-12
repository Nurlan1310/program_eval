from django.core.paginator import Paginator
from pathlib import Path
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import Group
from django.core.exceptions import PermissionDenied
from django.core.files.base import ContentFile
from django.http import FileResponse, Http404
from django.urls import reverse
from django.views.decorators.http import require_POST
from django.http import JsonResponse
import openpyxl, csv
from django.conf import settings

from evaluations.utils import log_action
from .ai_provider import generate_program_ai_report
from .analytics import build_program_ai_context, build_program_overview, extract_document_text
from .forms import ProgramForm, TopicForm, ImportForm, AIAnalyticsRunForm
from .models import ProgramAIAnalysisRun
from evaluations.models import Program, Topic, Criterion, TopicEvaluation, ProgramEvaluation, EvaluatorSession

import json
from datetime import datetime, timedelta
from django.db import transaction, connection
from django.utils.dateparse import parse_datetime

import io
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def is_subadmin(user):
    """Проверка, является ли пользователь субадмином"""
    if not user or not user.is_authenticated:
        return False
    # Создаем группу SubAdmins, если её нет
    group, created = Group.objects.get_or_create(name='SubAdmins')
    return user.is_superuser or user.groups.filter(name='SubAdmins').exists()


def manager_login(request):
    if request.GET.get('logout'):
        logout(request)
        return redirect('manager_login')

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        if not username or not password:
            error = 'Пожалуйста, введите логин и пароль.'
            return render(request, 'manager/login.html', {'error': error})
        
        user = authenticate(request, username=username, password=password)
        
        if not user:
            error = 'Неверные учётные данные. Проверьте логин и пароль.'
            return render(request, 'manager/login.html', {'error': error})
        
        if not user.is_active:
            error = 'Ваш аккаунт деактивирован. Обратитесь к администратору.'
            return render(request, 'manager/login.html', {'error': error})
        
        # Проверяем права доступа
        if not is_subadmin(user):
            # Создаем группу, если её нет
            group, created = Group.objects.get_or_create(name='SubAdmins')
            if created:
                error = f'Группа "SubAdmins" была создана. Пожалуйста, добавьте пользователя "{username}" в эту группу через админ-панель Django.'
            else:
                error = f'У пользователя "{username}" недостаточно прав. Добавьте пользователя в группу "SubAdmins" через админ-панель Django или используйте аккаунт суперадминистратора.'
            return render(request, 'manager/login.html', {'error': error})
        
        # Все проверки пройдены - выполняем вход
        login(request, user)
        # Логируем вход
        log_action(user, 'login', 'User', user.id, description=f'Вход в менеджерскую панель')
        return redirect('manager_index')
    
    return render(request, 'manager/login.html')


@login_required
def manager_index(request):
    if not is_subadmin(request.user): raise PermissionDenied
    programs = Program.objects.all()
    return render(request, 'manager/index.html', {'programs': programs})


@login_required
def program_add(request):
    if not is_subadmin(request.user): raise PermissionDenied
    if request.method == 'POST':
        form = ProgramForm(request.POST)
        if form.is_valid():
            program = form.save()
            log_action(request.user, 'create_program', 'Program', program.id, description=program.name)
            return redirect('manager_index')
    else:
        form = ProgramForm()
    return render(request, 'manager/program_add.html', {'form': form, 'breadcrumbs':[('Программы','/manager/'), ('Добавить','') ]})


@login_required
def program_detail(request, pk):
    if not is_subadmin(request.user): raise PermissionDenied
    program = get_object_or_404(Program, pk=pk)

    q = request.GET.get('q','').strip()
    topics = program.topics.all()
    if q:
        topics = topics.filter(name__icontains=q)
    paginator = Paginator(topics, 20)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)

    # Статистика
    total_topics = program.topics.count()
    # Считаем только тех оценщиков, которые действительно оценили хотя бы одну тему
    evaluators_count = TopicEvaluation.objects.filter(
        program_evaluation__program=program,
        completed_at__isnull=False
    ).values('program_evaluation__evaluator').distinct().count()
    completed_evaluations = TopicEvaluation.objects.filter(
        program_evaluation__program=program,
        completed_at__isnull=False
    ).count()
    if total_topics > 0:
        completion_percent = round((completed_evaluations / (total_topics * max(evaluators_count, 1))) * 100, 1)
    else:
        completion_percent = 0

    # Проверяем полноту закреплений оценщиков для модальных значений
    from evaluations.models import EvaluatorAssignment
    topics_all = program.topics.all()
    class_levels = set(t.class_level or "Без класса" for t in topics_all)
    assignments = EvaluatorAssignment.objects.filter(program=program)
    assigned_classes = set(a.class_level for a in assignments)
    missing_classes = class_levels - assigned_classes
    has_all_assignments = len(missing_classes) == 0

    return render(request, 'manager/program_detail.html', {
        'program': program,
        'page_obj': page_obj,
        'q': q,
        'total_topics': total_topics,
        'evaluators_count': evaluators_count,
        'completed_evaluations': completed_evaluations,
        'completion_percent': completion_percent,
        'has_all_assignments': has_all_assignments,
        'missing_classes': sorted(missing_classes) if missing_classes else [],
        'breadcrumbs': [('Программы','/manager/'), (program.name, '')]
    })

@login_required
def program_delete(request, pk):
    if not is_subadmin(request.user): raise PermissionDenied
    program = get_object_or_404(Program, pk=pk)
    name = program.name
    program.delete()
    log_action(request.user, 'delete_program', 'Program', pk, description=name)
    return redirect('manager_index')


@login_required
def topic_add(request, pk):
    if not is_subadmin(request.user): raise PermissionDenied
    program = get_object_or_404(Program, pk=pk)
    if request.method == 'POST':
        form = TopicForm(request.POST)
        if form.is_valid():
            topic = form.save(commit=False)
            topic.program = program
            topic.save()
            log_action(request.user, 'create_topic', 'Topic', topic.id, description=f'{topic.name} in {program.name}')
            return redirect('program_detail', pk=pk)
    else:
        form = TopicForm()
    return render(request, 'manager/topic_add.html', {'form': form, 'program': program, 'breadcrumbs':[('Программы','/manager/'), (program.name, '/manager/program/'+str(program.id)), ('Добавить тему','') ]})


@login_required
def topic_delete(request, pk):
    if not is_subadmin(request.user): raise PermissionDenied
    topic = get_object_or_404(Topic, pk=pk)
    program_id = topic.program.id
    name = topic.name
    topic.delete()
    log_action(request.user, 'delete_topic', 'Topic', pk, description=name)
    return redirect('program_detail', pk=program_id)


@login_required
def import_programs(request):
    if not is_subadmin(request.user):
        raise PermissionDenied

    # ----------------------------- #
    # STEP 1: Upload (no confirm)
    # ----------------------------- #
    if request.method == 'POST' and 'confirm' not in request.POST:
        form = ImportForm(request.POST, request.FILES)
        if not form.is_valid():
            return render(request, 'manager/import.html', {'form': form})

        file = request.FILES['file']
        rows = []

        # XLSX import
        if file.name.endswith('.xlsx'):
            wb = openpyxl.load_workbook(file)
            sheet = wb.active

            # expecting 3 columns: program, topic, level
            for p, t, lvl in sheet.iter_rows(min_row=2, values_only=True):
                if p:
                    rows.append((
                        str(p).strip(),
                        str(t).strip() if t else '',
                        str(lvl).strip() if lvl else ''
                    ))

        # CSV import
        else:
            text = file.read().decode('utf-8').splitlines()
            reader = csv.reader(text)
            next(reader, None)

            for p, t, lvl in reader:
                rows.append((
                    p.strip(),
                    t.strip() if t else '',
                    lvl.strip() if lvl else ''
                ))

        # Group by program
        programs = {}

        for p, t, lvl in rows:
            programs.setdefault(p, [])
            programs[p].append({
                "topic": t,
                "level": lvl
            })

        # Determine what exists
        analysis = []
        for pname, tlist in programs.items():
            exists = Program.objects.filter(name=pname).exists()

            analysis.append({
                "name": pname,
                "exists": exists,
                "topics": tlist
            })

        return render(request, 'manager/import_confirm.html', {
            'analysis': analysis,
            'json_data': json.dumps(programs),
        })

    # ----------------------------- #
    # STEP 2: Confirm import
    # ----------------------------- #
    if request.method == 'POST' and 'confirm' in request.POST:
        programs = json.loads(request.POST['data'])

        created_programs = 0
        created_topics = 0
        updated_topics = 0

        with transaction.atomic():
            for pname, topic_list in programs.items():

                program, prog_created = Program.objects.get_or_create(name=pname)
                if prog_created:
                    created_programs += 1

                order_counter = 1

                for item in topic_list:
                    tname = item["topic"]
                    lvl = item["level"] or "Без класса"

                    if not tname:
                        continue

                    topic, t_created = Topic.objects.get_or_create(
                        program=program,
                        name=tname,
                        class_level=lvl,
                        defaults={"order": order_counter}
                    )

                    if t_created:
                        created_topics += 1
                    else:
                        # update class_level if changed
                        if topic.class_level != lvl:
                            topic.class_level = lvl
                            topic.save(update_fields=["class_level"])
                            updated_topics += 1

                        # reorder
                        if topic.order != order_counter:
                            topic.order = order_counter
                            topic.save(update_fields=["order"])

                    order_counter += 1

        log_action(
            request.user,
            'import_programs',
            'Import',
            description=f'Programs: {created_programs}, '
                        f'New topics: {created_topics}, '
                        f'Updated topics: {updated_topics}'
        )

        return render(request, 'manager/import_done.html', {
            'created_programs': created_programs,
            'created_topics': created_topics,
            'updated_topics': updated_topics
        })

    # ----------------------------- #
    # STEP 0: Page with upload form
    # ----------------------------- #
    return render(request, 'manager/import.html', {'form': ImportForm()})


import datetime
import io
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter


def make_naive(dt):
    if dt is None: return None
    if dt.tzinfo is None: return dt
    return dt.replace(tzinfo=None)

def get_lastname(fullname):
    """Извлекает фамилию из полного имени (первое слово)."""
    if not fullname:
        return fullname
    return fullname.strip().split()[0]


def _calc_modal(values):
    """
    Возвращает (modal_value, degree_key, count_modal, total_answers)
    degree_key: conditional / acceptable / exact
    """
    if not values:
        return ("-", "conditional", 0, 0)
    freq = {}
    for v in values:
        freq[v] = freq.get(v, 0) + 1
    max_cnt = max(freq.values())
    # берём первое модальное значение с макс. частотой
    modal_val = next(k for k, v in freq.items() if v == max_cnt)
    total = len(values)

    if max_cnt == 1 and total > 1:
        degree = "conditional"
    elif max_cnt == 2:
        degree = "acceptable"
    else:
        degree = "exact"

    return (modal_val, degree, max_cnt, total)


def build_modal_block(ws, start_row, program, main_criteria, extra_criteria):
    """
    Лист модальностей в виде, похожем на Summary: блоки критериев по горизонтали,
    темы по вертикали, одна колонка на критерий с модальным значением.
    Использует закрепленных оценщиков из EvaluatorAssignment для каждого класса.
    """
    from evaluations.models import EvaluatorAssignment
    
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # более выразительные цвета
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")      # условное
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")   # допустимое
    green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")    # точное

    degree_fill = {
        "conditional": red_fill,
        "acceptable": yellow_fill,
        "exact": green_fill,
    }

    # Получаем закрепления оценщиков по классам
    assignments = EvaluatorAssignment.objects.filter(program=program).select_related(
        'evaluator1', 'evaluator2', 'evaluator3'
    )
    class_assignments = {a.class_level: a for a in assignments}

    # данные по ответам - используем только закрепленных оценщиков
    topic_evals = TopicEvaluation.objects.filter(
        program_evaluation__program=program,
        completed_at__isnull=False
    ).select_related("topic", "program_evaluation__evaluator").prefetch_related("answers__criterion")

    # Проверяем, что для всех классов есть закрепления
    topics = list(program.topics.all().order_by("order", "id"))
    class_levels = set(t.class_level or "Без класса" for t in topics)
    assigned_classes = set(class_assignments.keys())
    missing_classes = class_levels - assigned_classes
    
    if missing_classes:
        raise ValueError(
            f'Отсутствуют закрепления оценщиков для классов: {", ".join(sorted(missing_classes))}. '
            f'Необходимо закрепить оценщиков для всех классов перед экспортом модальных значений.'
        )
    
    answers_map = {}
    for te in topic_evals:
        topic = te.topic
        class_level = topic.class_level or "Без класса"
        assignment = class_assignments.get(class_level)
        
        # Должно быть закрепление для каждого класса (проверено выше)
        allowed_evaluators = {assignment.evaluator1, assignment.evaluator2, assignment.evaluator3}
        if te.program_evaluation.evaluator not in allowed_evaluators:
            continue
        
        for ans in te.answers.all():
            key = (te.topic_id, ans.criterion_id)
            val = ans.value if ans.value is not None else ("Yes" if ans.yes_no else "No")
            answers_map.setdefault(key, []).append(val)

    # Используем только основные критерии
    all_crit_obj = list(main_criteria)
    blocks = {}
    for c in all_crit_obj:
        block = c.block
        key = (block.order if block else 9999, block.name if block else "Без блока", block.id if block else None)
        blocks.setdefault(key, []).append(c)
    blocks_sorted = sorted(blocks.items(), key=lambda k: (k[0][0], k[0][1]))

    # Порядок критериев
    all_crit = []
    for _, crit_list in blocks_sorted:
        all_crit.extend(sorted(crit_list, key=lambda c: (c.order, c.id)))

    # Колонки вычислений для каждого блока: уровни 1-3, условные, проценты (4+4=8) + точные, допустимые, условные (3+3=6) = 14 колонок
    block_calc_cols = 14
    # Итоговые колонки: уровни 1-3, условные, проценты (4+4=8) + точные, допустимые, условные (3+3=6) = 14 колонок
    total_calc_cols = 14
    # Общее количество блоков
    num_blocks = len(blocks_sorted)
    total_cols = 1 + len(all_crit) + (num_blocks * block_calc_cols) + total_calc_cols

    row = start_row
    # Заголовок программы
    cell1 = ws.cell(row=row, column=1, value="Program")
    cell1.font = bold
    cell2 = ws.cell(row=row, column=2, value=program.name)
    cell2.font = bold
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=total_cols)
    row += 1

    # Заголовок блоков (объединяем до последней колонки вычислений блока)
    col = 2
    for (_order, block_name, _bid), crit_list in blocks_sorted:
        span = len(crit_list) + block_calc_cols  # Критерии + колонки вычислений
        cell = ws.cell(row=row, column=col, value=block_name)
        cell.alignment = center
        cell.font = bold
        ws.merge_cells(
            start_row=row,
            start_column=col,
            end_row=row,
            end_column=col + span - 1
        )
        col += span
    # Итоговые колонки
    cell = ws.cell(row=row, column=col, value="Итого")
    cell.alignment = center
    cell.font = bold
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + total_calc_cols - 1)
    row += 1

    # Заголовок критериев и вычислений (правильно под блоками)
    ws.cell(row=row, column=1, value="Topic").font = bold
    col = 2
    # Заголовки критериев и вычислений для каждого блока
    for (_order, block_name, _bid), crit_list in blocks_sorted:
        # Критерии блока
        for crit in sorted(crit_list, key=lambda c: (c.order, c.id)):
            cell = ws.cell(row=row, column=col, value=crit.name)
            cell.alignment = center
            cell.font = bold
            col += 1
        # Заголовки вычислений для блока
        ws.cell(row=row, column=col, value="Ур.1").font = bold; ws.cell(row=row, column=col).alignment = center; col += 1
        ws.cell(row=row, column=col, value="Ур.2").font = bold; ws.cell(row=row, column=col).alignment = center; col += 1
        ws.cell(row=row, column=col, value="Ур.3").font = bold; ws.cell(row=row, column=col).alignment = center; col += 1
        ws.cell(row=row, column=col, value="Усл.").font = bold; ws.cell(row=row, column=col).alignment = center; col += 1
        ws.cell(row=row, column=col, value="Ур.1%").font = bold; ws.cell(row=row, column=col).alignment = center; col += 1
        ws.cell(row=row, column=col, value="Ур.2%").font = bold; ws.cell(row=row, column=col).alignment = center; col += 1
        ws.cell(row=row, column=col, value="Ур.3%").font = bold; ws.cell(row=row, column=col).alignment = center; col += 1
        ws.cell(row=row, column=col, value="Усл.%").font = bold; ws.cell(row=row, column=col).alignment = center; col += 1
        ws.cell(row=row, column=col, value="Точн.").font = bold; ws.cell(row=row, column=col).alignment = center; col += 1
        ws.cell(row=row, column=col, value="Доп.").font = bold; ws.cell(row=row, column=col).alignment = center; col += 1
        ws.cell(row=row, column=col, value="Усл.").font = bold; ws.cell(row=row, column=col).alignment = center; col += 1
        ws.cell(row=row, column=col, value="Точн.%").font = bold; ws.cell(row=row, column=col).alignment = center; col += 1
        ws.cell(row=row, column=col, value="Доп.%").font = bold; ws.cell(row=row, column=col).alignment = center; col += 1
        ws.cell(row=row, column=col, value="Усл.%").font = bold; ws.cell(row=row, column=col).alignment = center; col += 1
    # Заголовки итоговых вычислений
    ws.cell(row=row, column=col, value="Ур.1").font = bold; ws.cell(row=row, column=col).alignment = center; col += 1
    ws.cell(row=row, column=col, value="Ур.2").font = bold; ws.cell(row=row, column=col).alignment = center; col += 1
    ws.cell(row=row, column=col, value="Ур.3").font = bold; ws.cell(row=row, column=col).alignment = center; col += 1
    ws.cell(row=row, column=col, value="Усл.").font = bold; ws.cell(row=row, column=col).alignment = center; col += 1
    ws.cell(row=row, column=col, value="Ур.1%").font = bold; ws.cell(row=row, column=col).alignment = center; col += 1
    ws.cell(row=row, column=col, value="Ур.2%").font = bold; ws.cell(row=row, column=col).alignment = center; col += 1
    ws.cell(row=row, column=col, value="Ур.3%").font = bold; ws.cell(row=row, column=col).alignment = center; col += 1
    ws.cell(row=row, column=col, value="Усл.%").font = bold; ws.cell(row=row, column=col).alignment = center; col += 1
    ws.cell(row=row, column=col, value="Точн.").font = bold; ws.cell(row=row, column=col).alignment = center; col += 1
    ws.cell(row=row, column=col, value="Доп.").font = bold; ws.cell(row=row, column=col).alignment = center; col += 1
    ws.cell(row=row, column=col, value="Усл.").font = bold; ws.cell(row=row, column=col).alignment = center; col += 1
    ws.cell(row=row, column=col, value="Точн.%").font = bold; ws.cell(row=row, column=col).alignment = center; col += 1
    ws.cell(row=row, column=col, value="Доп.%").font = bold; ws.cell(row=row, column=col).alignment = center; col += 1
    ws.cell(row=row, column=col, value="Усл.%").font = bold; ws.cell(row=row, column=col).alignment = center
    row += 1

    # Итоги счётчиков
    degree_totals = {"conditional": 0, "acceptable": 0, "exact": 0}
    value_totals = {}
    block_degree_totals = {k[1]: {"conditional": 0, "acceptable": 0, "exact": 0} for k, _ in blocks_sorted}
    # Итоги по критериям (по столбцам)
    crit_degree_totals = {crit.id: {"conditional": 0, "acceptable": 0, "exact": 0} for crit in all_crit}
    
    # Итоговые данные по всем классам для расчета по уровням
    total_modal_values = {crit.id: [] for crit in all_crit}  # Только допустимые и точные
    total_conditional_counts = {crit.id: 0 for crit in all_crit}  # Условные
    total_degree_counts = {crit.id: {"conditional": 0, "acceptable": 0, "exact": 0} for crit in all_crit}  # Итоговые степени модальности

    # Темы по классам
    grouped_topics = {}
    for t in topics:
        cls = t.class_level or "Без класса"
        grouped_topics.setdefault(cls, []).append(t)

    for cls_name in sorted(grouped_topics.keys()):
        # строка класса
        cell = ws.cell(row=row, column=1, value=f"Class: {cls_name}")
        cell.font = bold
        cell.alignment = center
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
        row += 1

        # Собираем модальные значения по темам класса для расчета по уровням
        # Только для допустимых и точных (желтые и зеленые)
        class_modal_values = {crit.id: [] for crit in all_crit}
        class_conditional_counts = {crit.id: 0 for crit in all_crit}  # Количество условных для каждого критерия
        class_degree_counts = {crit.id: {"conditional": 0, "acceptable": 0, "exact": 0} for crit in all_crit}  # Подсчет степеней модальности
        
        for topic in grouped_topics[cls_name]:
            ws.cell(row=row, column=1, value=topic.name)
            col = 2
            # подсчёт по теме
            topic_degree = {"conditional": 0, "acceptable": 0, "exact": 0}
            # Данные для вычислений по теме
            topic_modal_values = []  # Модальные значения для темы (только допустимые и точные)
            topic_conditional_count = 0  # Количество условных для темы
            topic_block_data = {}  # Данные по блокам для темы
            topic_block_degrees = {}  # Данные по степеням модальности для каждого блока
            
            # Создаем маппинг критериев к блокам
            crit_to_block = {}
            for (_order, block_name, _bid), crit_list in blocks_sorted:
                for crit in sorted(crit_list, key=lambda c: (c.order, c.id)):
                    crit_to_block[crit.id] = block_name
            
            for crit in all_crit:
                vals = answers_map.get((topic.id, crit.id), [])
                modal_val, degree, _cnt, _total = _calc_modal(vals)

                ws.cell(row=row, column=col, value=modal_val)
                ws.cell(row=row, column=col).alignment = center
                ws.cell(row=row, column=col).fill = degree_fill.get(degree, red_fill)

                # Сохраняем модальное значение для расчета по уровням класса
                # Только для допустимых и точных (acceptable, exact)
                if degree in ["acceptable", "exact"] and modal_val != "-" and isinstance(modal_val, (int, str)):
                    try:
                        val_int = int(modal_val)
                        if val_int in [1, 2, 3]:
                            class_modal_values[crit.id].append(val_int)
                            # Также добавляем в общий итог
                            total_modal_values[crit.id].append(val_int)
                            # Для вычислений по теме
                            topic_modal_values.append(val_int)
                    except (ValueError, TypeError):
                        pass
                
                # Считаем условные отдельно
                if degree == "conditional":
                    class_conditional_counts[crit.id] += 1
                    # Также добавляем в общий итог
                    total_conditional_counts[crit.id] += 1
                    # Для вычислений по теме
                    topic_conditional_count += 1
                
                # Подсчитываем степени модальности для класса
                class_degree_counts[crit.id][degree] += 1
                # Также добавляем в общий итог
                total_degree_counts[crit.id][degree] += 1
                
                # Данные для вычислений по блоку
                block_name = crit_to_block.get(crit.id, "Без блока")
                if block_name not in topic_block_data:
                    topic_block_data[block_name] = {"modal_values": [], "conditional": 0}
                    topic_block_degrees[block_name] = {"conditional": 0, "acceptable": 0, "exact": 0}
                if degree in ["acceptable", "exact"] and modal_val != "-" and isinstance(modal_val, (int, str)):
                    try:
                        val_int = int(modal_val)
                        if val_int in [1, 2, 3]:
                            topic_block_data[block_name]["modal_values"].append(val_int)
                    except (ValueError, TypeError):
                        pass
                if degree == "conditional":
                    topic_block_data[block_name]["conditional"] += 1
                # Подсчитываем степени модальности для блока
                topic_block_degrees[block_name][degree] += 1

                degree_totals[degree] += 1
                topic_degree[degree] += 1
                crit_degree_totals[crit.id][degree] += 1
                value_totals[modal_val] = value_totals.get(modal_val, 0) + 1
                bname = crit.block.name if crit.block else "Без блока"
                block_degree_totals.setdefault(bname, {"conditional": 0, "acceptable": 0, "exact": 0})
                block_degree_totals[bname][degree] += 1

                col += 1
                
                # После последнего критерия в блоке добавляем вычисления для блока
                block_name = crit_to_block.get(crit.id, "Без блока")
                # Проверяем, является ли это последним критерием в блоке
                block_crits = [c for c in all_crit if crit_to_block.get(c.id) == block_name]
                if crit == block_crits[-1]:  # Последний критерий в блоке
                    block_data = topic_block_data.get(block_name, {"modal_values": [], "conditional": 0})
                    block_degrees = topic_block_degrees.get(block_name, {"conditional": 0, "acceptable": 0, "exact": 0})
                    num_crits_in_block = len(block_crits)
                    # Уровни 1-3
                    for level in [1, 2, 3]:
                        count = block_data["modal_values"].count(level)
                        ws.cell(row=row, column=col, value=count).alignment = center
                        col += 1
                    # Условные
                    ws.cell(row=row, column=col, value=block_data["conditional"]).alignment = center
                    col += 1
                    # Проценты для уровней
                    for level in [1, 2, 3]:
                        count = block_data["modal_values"].count(level)
                        percentage = (count / num_crits_in_block * 100) if num_crits_in_block > 0 else 0
                        ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
                        col += 1
                    # Процент для условных
                    percentage = (block_data["conditional"] / num_crits_in_block * 100) if num_crits_in_block > 0 else 0
                    ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
                    col += 1
                    # Точные, допустимые, условные (числа)
                    ws.cell(row=row, column=col, value=block_degrees["exact"]).fill = green_fill
                    ws.cell(row=row, column=col).alignment = center
                    col += 1
                    ws.cell(row=row, column=col, value=block_degrees["acceptable"]).fill = yellow_fill
                    ws.cell(row=row, column=col).alignment = center
                    col += 1
                    ws.cell(row=row, column=col, value=block_degrees["conditional"]).fill = red_fill
                    ws.cell(row=row, column=col).alignment = center
                    col += 1
                    # Точные %, допустимые %, условные % (проценты)
                    for degree_key in ["exact", "acceptable", "conditional"]:
                        count = block_degrees[degree_key]
                        percentage = (count / num_crits_in_block * 100) if num_crits_in_block > 0 else 0
                        ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
                        col += 1

            # Итоговые вычисления для темы (по всем критериям)
            num_crits = len(all_crit)
            # Уровни 1-3
            for level in [1, 2, 3]:
                count = topic_modal_values.count(level)
                ws.cell(row=row, column=col, value=count).alignment = center
                col += 1
            # Условные
            ws.cell(row=row, column=col, value=topic_conditional_count).alignment = center
            col += 1
            # Проценты для уровней
            for level in [1, 2, 3]:
                count = topic_modal_values.count(level)
                percentage = (count / num_crits * 100) if num_crits > 0 else 0
                ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
                col += 1
            # Процент для условных
            percentage = (topic_conditional_count / num_crits * 100) if num_crits > 0 else 0
            ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
            col += 1
            # Точные, допустимые, условные (числа)
            ws.cell(row=row, column=col, value=topic_degree["exact"]).fill = green_fill
            ws.cell(row=row, column=col).alignment = center
            col += 1
            ws.cell(row=row, column=col, value=topic_degree["acceptable"]).fill = yellow_fill
            ws.cell(row=row, column=col).alignment = center
            col += 1
            ws.cell(row=row, column=col, value=topic_degree["conditional"]).fill = red_fill
            ws.cell(row=row, column=col).alignment = center
            col += 1
            # Точные %, допустимые %, условные % (проценты)
            for degree_key in ["exact", "acceptable", "conditional"]:
                count = topic_degree[degree_key]
                percentage = (count / num_crits * 100) if num_crits > 0 else 0
                ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
                col += 1

            row += 1

        # Строки расчета по уровням для этого класса (только для допустимых и точных)
        # Создаем маппинг критериев к блокам
        crit_to_block = {}
        for (_order, block_name, _bid), crit_list in blocks_sorted:
            for crit in sorted(crit_list, key=lambda c: (c.order, c.id)):
                crit_to_block[crit.id] = block_name
        
        total_topics_in_class = len(grouped_topics[cls_name])
        
        for level in [1, 2, 3]:
            ws.cell(row=row, column=1, value=f"Уровень {level}").font = bold
            col = 2
            for crit in all_crit:
                count = class_modal_values[crit.id].count(level)
                ws.cell(row=row, column=col, value=count)
                ws.cell(row=row, column=col).alignment = center
                col += 1
                
                # После последнего критерия в блоке добавляем вычисления для блока
                block_name = crit_to_block.get(crit.id, "Без блока")
                block_crits = [c for c in all_crit if crit_to_block.get(c.id) == block_name]
                if crit == block_crits[-1]:  # Последний критерий в блоке
                    # Собираем данные по блоку для этого уровня
                    block_modal_values = []
                    for c in block_crits:
                        block_modal_values.extend([v for v in class_modal_values[c.id] if v == level])
                    block_count = len(block_modal_values)
                    num_crits_in_block = len(block_crits)
                    # Уровни 1-3
                    ws.cell(row=row, column=col, value=block_count).alignment = center
                    col += 1
                    # Остальные уровни (пустые для этой строки)
                    for _ in range(2):
                        ws.cell(row=row, column=col, value="").alignment = center
                        col += 1
                    # Условные (пустые для этой строки)
                    ws.cell(row=row, column=col, value="").alignment = center
                    col += 1
                    # Проценты для уровней
                    percentage = (block_count / (num_crits_in_block * total_topics_in_class) * 100) if (num_crits_in_block * total_topics_in_class) > 0 else 0
                    ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
                    col += 1
                    # Остальные проценты (пустые)
                    for _ in range(3):
                        ws.cell(row=row, column=col, value="").alignment = center
                        col += 1
                    # Точные, допустимые, условные (пустые для этой строки)
                    for _ in range(3):
                        ws.cell(row=row, column=col, value="").alignment = center
                        col += 1
                    # Точные %, допустимые %, условные % (пустые для этой строки)
                    for _ in range(3):
                        ws.cell(row=row, column=col, value="").alignment = center
                        col += 1
            
            # Итоговые вычисления для этого уровня
            total_level_count = sum(class_modal_values[crit.id].count(level) for crit in all_crit)
            ws.cell(row=row, column=col, value=total_level_count).alignment = center
            col += 1
            # Остальные уровни (пустые)
            for _ in range(2):
                ws.cell(row=row, column=col, value="").alignment = center
                col += 1
            # Условные (пустые)
            ws.cell(row=row, column=col, value="").alignment = center
            col += 1
            # Проценты
            total_possible = len(all_crit) * total_topics_in_class
            percentage = (total_level_count / total_possible * 100) if total_possible > 0 else 0
            ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
            col += 1
            # Остальные проценты (пустые)
            for _ in range(3):
                ws.cell(row=row, column=col, value="").alignment = center
                col += 1
            # Точные, допустимые, условные (пустые для этой строки)
            for _ in range(3):
                ws.cell(row=row, column=col, value="").alignment = center
                col += 1
            # Точные %, допустимые %, условные % (пустые для этой строки)
            for _ in range(3):
                ws.cell(row=row, column=col, value="").alignment = center
                col += 1
            
            row += 1
        
        # Строка "Условные" для этого класса
        ws.cell(row=row, column=1, value="Условные").font = bold
        col = 2
        for crit in all_crit:
            count = class_conditional_counts[crit.id]
            ws.cell(row=row, column=col, value=count)
            ws.cell(row=row, column=col).alignment = center
            col += 1
            
            # После последнего критерия в блоке добавляем вычисления для блока
            block_name = crit_to_block.get(crit.id, "Без блока")
            block_crits = [c for c in all_crit if crit_to_block.get(c.id) == block_name]
            if crit == block_crits[-1]:  # Последний критерий в блоке
                # Собираем условные по блоку
                block_conditional = sum(class_conditional_counts[c.id] for c in block_crits)
                num_crits_in_block = len(block_crits)
                # Уровни (пустые для этой строки)
                for _ in range(3):
                    ws.cell(row=row, column=col, value="").alignment = center
                    col += 1
                # Условные
                ws.cell(row=row, column=col, value=block_conditional).alignment = center
                col += 1
                # Проценты для уровней (пустые)
                for _ in range(3):
                    ws.cell(row=row, column=col, value="").alignment = center
                    col += 1
                # Процент для условных
                percentage = (block_conditional / (num_crits_in_block * total_topics_in_class) * 100) if (num_crits_in_block * total_topics_in_class) > 0 else 0
                ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
                col += 1
                # Точные, допустимые, условные (пустые для этой строки)
                for _ in range(3):
                    ws.cell(row=row, column=col, value="").alignment = center
                    col += 1
                # Точные %, допустимые %, условные % (пустые для этой строки)
                for _ in range(3):
                    ws.cell(row=row, column=col, value="").alignment = center
                    col += 1
        
        # Итоговые вычисления для условных
        total_conditional = sum(class_conditional_counts[crit.id] for crit in all_crit)
        # Уровни (пустые)
        for _ in range(3):
            ws.cell(row=row, column=col, value="").alignment = center
            col += 1
        # Условные
        ws.cell(row=row, column=col, value=total_conditional).alignment = center
        col += 1
        # Проценты для уровней (пустые)
        for _ in range(3):
            ws.cell(row=row, column=col, value="").alignment = center
            col += 1
        # Процент для условных
        total_possible = len(all_crit) * total_topics_in_class
        percentage = (total_conditional / total_possible * 100) if total_possible > 0 else 0
        ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
        col += 1
        # Точные, допустимые, условные (пустые для этой строки)
        for _ in range(3):
            ws.cell(row=row, column=col, value="").alignment = center
            col += 1
        # Точные %, допустимые %, условные % (пустые для этой строки)
        for _ in range(3):
            ws.cell(row=row, column=col, value="").alignment = center
            col += 1
        
        row += 1
        
        # Строки с процентами для этого класса
        if total_topics_in_class > 0:
            # Проценты для уровней
            for level in [1, 2, 3]:
                ws.cell(row=row, column=1, value=f"Уровень {level} %").font = bold
                col = 2
                for crit in all_crit:
                    count = class_modal_values[crit.id].count(level)
                    percentage = (count / total_topics_in_class * 100) if total_topics_in_class > 0 else 0
                    ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%")
                    ws.cell(row=row, column=col).alignment = center
                    col += 1
                    
                    # После последнего критерия в блоке добавляем вычисления для блока
                    block_name = crit_to_block.get(crit.id, "Без блока")
                    block_crits = [c for c in all_crit if crit_to_block.get(c.id) == block_name]
                    if crit == block_crits[-1]:  # Последний критерий в блоке
                        # Собираем данные по блоку для этого уровня
                        block_modal_values = []
                        for c in block_crits:
                            block_modal_values.extend([v for v in class_modal_values[c.id] if v == level])
                        block_count = len(block_modal_values)
                        num_crits_in_block = len(block_crits)
                        # Процент для этого уровня блока
                        percentage = (block_count / (num_crits_in_block * total_topics_in_class) * 100) if (num_crits_in_block * total_topics_in_class) > 0 else 0
                        ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
                        col += 1
                        # Остальные уровни (пустые)
                        for _ in range(2):
                            ws.cell(row=row, column=col, value="").alignment = center
                            col += 1
                        # Условные (пустые)
                        ws.cell(row=row, column=col, value="").alignment = center
                        col += 1
                        # Проценты для уровней (пустые, так как это уже проценты)
                        for _ in range(3):
                            ws.cell(row=row, column=col, value="").alignment = center
                            col += 1
                        # Процент для условных (пустой)
                        ws.cell(row=row, column=col, value="").alignment = center
                        col += 1
                        # Точные, допустимые, условные (пустые для этой строки)
                        for _ in range(3):
                            ws.cell(row=row, column=col, value="").alignment = center
                            col += 1
                        # Точные %, допустимые %, условные % (пустые для этой строки)
                        for _ in range(3):
                            ws.cell(row=row, column=col, value="").alignment = center
                            col += 1
                
                # Итоговые вычисления для этого уровня
                total_level_count = sum(class_modal_values[crit.id].count(level) for crit in all_crit)
                total_possible = len(all_crit) * total_topics_in_class
                percentage = (total_level_count / total_possible * 100) if total_possible > 0 else 0
                ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
                col += 1
                # Остальные уровни (пустые)
                for _ in range(2):
                    ws.cell(row=row, column=col, value="").alignment = center
                    col += 1
                # Условные (пустые)
                ws.cell(row=row, column=col, value="").alignment = center
                col += 1
                # Проценты (пустые)
                for _ in range(4):
                    ws.cell(row=row, column=col, value="").alignment = center
                    col += 1
                # Точные, допустимые, условные (пустые для этой строки)
                for _ in range(3):
                    ws.cell(row=row, column=col, value="").alignment = center
                    col += 1
                # Точные %, допустимые %, условные % (пустые для этой строки)
                for _ in range(3):
                    ws.cell(row=row, column=col, value="").alignment = center
                    col += 1
                
                row += 1
            
            # Процент для условных
            ws.cell(row=row, column=1, value="Условные %").font = bold
            col = 2
            for crit in all_crit:
                count = class_conditional_counts[crit.id]
                percentage = (count / total_topics_in_class * 100) if total_topics_in_class > 0 else 0
                ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%")
                ws.cell(row=row, column=col).alignment = center
                col += 1
                
                # После последнего критерия в блоке добавляем вычисления для блока
                block_name = crit_to_block.get(crit.id, "Без блока")
                block_crits = [c for c in all_crit if crit_to_block.get(c.id) == block_name]
                if crit == block_crits[-1]:  # Последний критерий в блоке
                    # Собираем условные по блоку
                    block_conditional = sum(class_conditional_counts[c.id] for c in block_crits)
                    num_crits_in_block = len(block_crits)
                    # Уровни (пустые)
                    for _ in range(3):
                        ws.cell(row=row, column=col, value="").alignment = center
                        col += 1
                    # Условные (пустые, так как это уже проценты)
                    ws.cell(row=row, column=col, value="").alignment = center
                    col += 1
                    # Проценты для уровней (пустые)
                    for _ in range(3):
                        ws.cell(row=row, column=col, value="").alignment = center
                        col += 1
                    # Процент для условных
                    percentage = (block_conditional / (num_crits_in_block * total_topics_in_class) * 100) if (num_crits_in_block * total_topics_in_class) > 0 else 0
                    ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
                    col += 1
                    # Точные, допустимые, условные (пустые для этой строки)
                    for _ in range(3):
                        ws.cell(row=row, column=col, value="").alignment = center
                        col += 1
                    # Точные %, допустимые %, условные % (пустые для этой строки)
                    for _ in range(3):
                        ws.cell(row=row, column=col, value="").alignment = center
                        col += 1
            
            # Итоговые вычисления для условных
            total_conditional = sum(class_conditional_counts[crit.id] for crit in all_crit)
            # Уровни (пустые)
            for _ in range(3):
                ws.cell(row=row, column=col, value="").alignment = center
                col += 1
            # Условные (пустые)
            ws.cell(row=row, column=col, value="").alignment = center
            col += 1
            # Проценты для уровней (пустые)
            for _ in range(3):
                ws.cell(row=row, column=col, value="").alignment = center
                col += 1
            # Процент для условных
            total_conditional = sum(class_conditional_counts[crit.id] for crit in all_crit)
            total_possible = len(all_crit) * total_topics_in_class
            percentage = (total_conditional / total_possible * 100) if total_possible > 0 else 0
            ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
            col += 1
            # Точные, допустимые, условные (пустые для этой строки)
            for _ in range(3):
                ws.cell(row=row, column=col, value="").alignment = center
                col += 1
            # Точные %, допустимые %, условные % (пустые для этой строки)
            for _ in range(3):
                ws.cell(row=row, column=col, value="").alignment = center
                col += 1
            
            row += 1
            
            # Подсчет точных, допустимых и условных значений для этого класса
            for degree_key, label, fill_color in [("exact", "Точные", green_fill), ("acceptable", "Допустимые", yellow_fill), ("conditional", "Условные", red_fill)]:
                ws.cell(row=row, column=1, value=label).font = bold
                ws.cell(row=row, column=1).fill = fill_color  # Закрашиваем ячейку названия
                col = 2
                for crit in all_crit:
                    count = class_degree_counts[crit.id][degree_key]
                    ws.cell(row=row, column=col, value=count)
                    ws.cell(row=row, column=col).fill = fill_color
                    ws.cell(row=row, column=col).alignment = center
                    col += 1
                    
                    # После последнего критерия в блоке добавляем вычисления для блока
                    block_name = crit_to_block.get(crit.id, "Без блока")
                    block_crits = [c for c in all_crit if crit_to_block.get(c.id) == block_name]
                    if crit == block_crits[-1]:  # Последний критерий в блоке
                        # Собираем данные по блоку для этой степени
                        block_count = sum(class_degree_counts[c.id][degree_key] for c in block_crits)
                        num_crits_in_block = len(block_crits)
                        # Уровни (пустые для этой строки)
                        for _ in range(3):
                            ws.cell(row=row, column=col, value="").alignment = center
                            col += 1
                        # Условные (пустые для этой строки, кроме строки "Условные")
                        if degree_key == "conditional":
                            ws.cell(row=row, column=col, value=block_count).fill = fill_color
                            ws.cell(row=row, column=col).alignment = center
                        else:
                            ws.cell(row=row, column=col, value="").alignment = center
                        col += 1
                        # Проценты для уровней (пустые)
                        for _ in range(3):
                            ws.cell(row=row, column=col, value="").alignment = center
                            col += 1
                        # Процент для условных (пустой, кроме строки "Условные")
                        if degree_key == "conditional":
                            percentage = (block_count / (num_crits_in_block * total_topics_in_class) * 100) if (num_crits_in_block * total_topics_in_class) > 0 else 0
                            ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
                        else:
                            ws.cell(row=row, column=col, value="").alignment = center
                        col += 1
                        # Точные, допустимые, условные (числа) - заполняем только соответствующую колонку
                        if degree_key == "exact":
                            ws.cell(row=row, column=col, value=block_count).fill = fill_color
                            ws.cell(row=row, column=col).alignment = center
                        else:
                            ws.cell(row=row, column=col, value="").alignment = center
                        col += 1
                        if degree_key == "acceptable":
                            ws.cell(row=row, column=col, value=block_count).fill = fill_color
                            ws.cell(row=row, column=col).alignment = center
                        else:
                            ws.cell(row=row, column=col, value="").alignment = center
                        col += 1
                        if degree_key == "conditional":
                            ws.cell(row=row, column=col, value=block_count).fill = fill_color
                            ws.cell(row=row, column=col).alignment = center
                        else:
                            ws.cell(row=row, column=col, value="").alignment = center
                        col += 1
                        # Точные %, допустимые %, условные % (проценты) - заполняем только соответствующую колонку
                        if degree_key == "exact":
                            percentage = (block_count / (num_crits_in_block * total_topics_in_class) * 100) if (num_crits_in_block * total_topics_in_class) > 0 else 0
                            ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
                        else:
                            ws.cell(row=row, column=col, value="").alignment = center
                        col += 1
                        if degree_key == "acceptable":
                            percentage = (block_count / (num_crits_in_block * total_topics_in_class) * 100) if (num_crits_in_block * total_topics_in_class) > 0 else 0
                            ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
                        else:
                            ws.cell(row=row, column=col, value="").alignment = center
                        col += 1
                        if degree_key == "conditional":
                            percentage = (block_count / (num_crits_in_block * total_topics_in_class) * 100) if (num_crits_in_block * total_topics_in_class) > 0 else 0
                            ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
                        else:
                            ws.cell(row=row, column=col, value="").alignment = center
                        col += 1
                
                # Итоговые вычисления для этой степени
                total_count = sum(class_degree_counts[crit.id][degree_key] for crit in all_crit)
                # Уровни (пустые)
                for _ in range(3):
                    ws.cell(row=row, column=col, value="").alignment = center
                    col += 1
                # Условные (пустые, кроме строки "Условные")
                if degree_key == "conditional":
                    ws.cell(row=row, column=col, value=total_count).fill = fill_color
                    ws.cell(row=row, column=col).alignment = center
                else:
                    ws.cell(row=row, column=col, value="").alignment = center
                col += 1
                # Проценты для уровней (пустые)
                for _ in range(3):
                    ws.cell(row=row, column=col, value="").alignment = center
                    col += 1
                # Процент для условных (пустой, кроме строки "Условные")
                if degree_key == "conditional":
                    total_possible = len(all_crit) * total_topics_in_class
                    percentage = (total_count / total_possible * 100) if total_possible > 0 else 0
                    ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
                else:
                    ws.cell(row=row, column=col, value="").alignment = center
                col += 1
                # Точные, допустимые, условные (числа) - заполняем только соответствующую колонку
                if degree_key == "exact":
                    ws.cell(row=row, column=col, value=total_count).fill = fill_color
                    ws.cell(row=row, column=col).alignment = center
                else:
                    ws.cell(row=row, column=col, value="").alignment = center
                col += 1
                if degree_key == "acceptable":
                    ws.cell(row=row, column=col, value=total_count).fill = fill_color
                    ws.cell(row=row, column=col).alignment = center
                else:
                    ws.cell(row=row, column=col, value="").alignment = center
                col += 1
                if degree_key == "conditional":
                    ws.cell(row=row, column=col, value=total_count).fill = fill_color
                    ws.cell(row=row, column=col).alignment = center
                else:
                    ws.cell(row=row, column=col, value="").alignment = center
                col += 1
                # Точные %, допустимые %, условные % (проценты) - заполняем только соответствующую колонку
                if degree_key == "exact":
                    total_possible = len(all_crit) * total_topics_in_class
                    percentage = (total_count / total_possible * 100) if total_possible > 0 else 0
                    ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
                else:
                    ws.cell(row=row, column=col, value="").alignment = center
                col += 1
                if degree_key == "acceptable":
                    total_possible = len(all_crit) * total_topics_in_class
                    percentage = (total_count / total_possible * 100) if total_possible > 0 else 0
                    ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
                else:
                    ws.cell(row=row, column=col, value="").alignment = center
                col += 1
                if degree_key == "conditional":
                    total_possible = len(all_crit) * total_topics_in_class
                    percentage = (total_count / total_possible * 100) if total_possible > 0 else 0
                    ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
                else:
                    ws.cell(row=row, column=col, value="").alignment = center
                
                row += 1
            
            # Проценты для точных, допустимых и условных
            for degree_key, label in [("exact", "Точные %"), ("acceptable", "Допустимые %"), ("conditional", "Условные %")]:
                ws.cell(row=row, column=1, value=label).font = bold
                col = 2
                for crit in all_crit:
                    count = class_degree_counts[crit.id][degree_key]
                    percentage = (count / total_topics_in_class * 100) if total_topics_in_class > 0 else 0
                    ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%")
                    ws.cell(row=row, column=col).alignment = center
                    col += 1
                    
                    # После последнего критерия в блоке добавляем вычисления для блока
                    block_name = crit_to_block.get(crit.id, "Без блока")
                    block_crits = [c for c in all_crit if crit_to_block.get(c.id) == block_name]
                    if crit == block_crits[-1]:  # Последний критерий в блоке
                        # Собираем данные по блоку для этой степени
                        block_count = sum(class_degree_counts[c.id][degree_key] for c in block_crits)
                        num_crits_in_block = len(block_crits)
                        # Уровни (пустые)
                        for _ in range(3):
                            ws.cell(row=row, column=col, value="").alignment = center
                            col += 1
                        # Условные (пустые, так как это уже проценты)
                        ws.cell(row=row, column=col, value="").alignment = center
                        col += 1
                        # Проценты для уровней (пустые, так как это уже проценты)
                        for _ in range(3):
                            ws.cell(row=row, column=col, value="").alignment = center
                            col += 1
                        # Процент для условных
                        if degree_key == "conditional":
                            percentage = (block_count / (num_crits_in_block * total_topics_in_class) * 100) if (num_crits_in_block * total_topics_in_class) > 0 else 0
                            ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
                        else:
                            ws.cell(row=row, column=col, value="").alignment = center
                        col += 1
                        # Точные %, допустимые %, условные % (проценты) - заполняем только соответствующую колонку
                        if degree_key == "exact":
                            percentage = (block_count / (num_crits_in_block * total_topics_in_class) * 100) if (num_crits_in_block * total_topics_in_class) > 0 else 0
                            ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
                        else:
                            ws.cell(row=row, column=col, value="").alignment = center
                        col += 1
                        if degree_key == "acceptable":
                            percentage = (block_count / (num_crits_in_block * total_topics_in_class) * 100) if (num_crits_in_block * total_topics_in_class) > 0 else 0
                            ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
                        else:
                            ws.cell(row=row, column=col, value="").alignment = center
                        col += 1
                        if degree_key == "conditional":
                            percentage = (block_count / (num_crits_in_block * total_topics_in_class) * 100) if (num_crits_in_block * total_topics_in_class) > 0 else 0
                            ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
                        else:
                            ws.cell(row=row, column=col, value="").alignment = center
                        col += 1
                        # Точные, допустимые, условные (пустые, так как это уже проценты)
                        for _ in range(3):
                            ws.cell(row=row, column=col, value="").alignment = center
                            col += 1
                
                # Итоговые вычисления для этой степени
                total_count = sum(class_degree_counts[crit.id][degree_key] for crit in all_crit)
                # Уровни (пустые)
                for _ in range(3):
                    ws.cell(row=row, column=col, value="").alignment = center
                    col += 1
                # Условные (пустые)
                ws.cell(row=row, column=col, value="").alignment = center
                col += 1
                # Проценты для уровней (пустые)
                for _ in range(3):
                    ws.cell(row=row, column=col, value="").alignment = center
                    col += 1
                # Процент для условных
                if degree_key == "conditional":
                    total_possible = len(all_crit) * total_topics_in_class
                    percentage = (total_count / total_possible * 100) if total_possible > 0 else 0
                    ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
                else:
                    ws.cell(row=row, column=col, value="").alignment = center
                col += 1
                # Точные %, допустимые %, условные % (проценты) - заполняем только соответствующую колонку
                if degree_key == "exact":
                    total_possible = len(all_crit) * total_topics_in_class
                    percentage = (total_count / total_possible * 100) if total_possible > 0 else 0
                    ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
                else:
                    ws.cell(row=row, column=col, value="").alignment = center
                col += 1
                if degree_key == "acceptable":
                    total_possible = len(all_crit) * total_topics_in_class
                    percentage = (total_count / total_possible * 100) if total_possible > 0 else 0
                    ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
                else:
                    ws.cell(row=row, column=col, value="").alignment = center
                col += 1
                if degree_key == "conditional":
                    total_possible = len(all_crit) * total_topics_in_class
                    percentage = (total_count / total_possible * 100) if total_possible > 0 else 0
                    ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
                else:
                    ws.cell(row=row, column=col, value="").alignment = center
                col += 1
                # Точные, допустимые, условные (пустые, так как это уже проценты)
                for _ in range(3):
                    ws.cell(row=row, column=col, value="").alignment = center
                    col += 1
                
                row += 1

    # Итоговый подсчет по всем классам вместе
    cell = ws.cell(row=row, column=1, value="Итого по всем классам")
    cell.font = Font(bold=True, size=12)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
    row += 1
    
    # Строки расчета по уровням для всех классов (только допустимые и точные)
    # Создаем маппинг критериев к блокам
    crit_to_block = {}
    for (_order, block_name, _bid), crit_list in blocks_sorted:
        for crit in sorted(crit_list, key=lambda c: (c.order, c.id)):
            crit_to_block[crit.id] = block_name
    
    total_topics_all = sum(len(grouped_topics[cls]) for cls in grouped_topics.keys())
    
    for level in [1, 2, 3]:
        ws.cell(row=row, column=1, value=f"Уровень {level}").font = bold
        col = 2
        for crit in all_crit:
            count = total_modal_values[crit.id].count(level)
            ws.cell(row=row, column=col, value=count)
            ws.cell(row=row, column=col).alignment = center
            col += 1
            
            # После последнего критерия в блоке добавляем вычисления для блока
            block_name = crit_to_block.get(crit.id, "Без блока")
            block_crits = [c for c in all_crit if crit_to_block.get(c.id) == block_name]
            if crit == block_crits[-1]:  # Последний критерий в блоке
                # Собираем данные по блоку для этого уровня
                block_modal_values = []
                for c in block_crits:
                    block_modal_values.extend([v for v in total_modal_values[c.id] if v == level])
                block_count = len(block_modal_values)
                num_crits_in_block = len(block_crits)
                # Уровни 1-3
                ws.cell(row=row, column=col, value=block_count).alignment = center
                col += 1
                # Остальные уровни (пустые для этой строки)
                for _ in range(2):
                    ws.cell(row=row, column=col, value="").alignment = center
                    col += 1
                # Условные (пустые для этой строки)
                ws.cell(row=row, column=col, value="").alignment = center
                col += 1
                # Проценты для уровней
                percentage = (block_count / (num_crits_in_block * total_topics_all) * 100) if (num_crits_in_block * total_topics_all) > 0 else 0
                ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
                col += 1
                # Остальные проценты (пустые)
                for _ in range(3):
                    ws.cell(row=row, column=col, value="").alignment = center
                    col += 1
                # Точные, допустимые, условные (пустые для этой строки)
                for _ in range(3):
                    ws.cell(row=row, column=col, value="").alignment = center
                    col += 1
                # Точные %, допустимые %, условные % (пустые для этой строки)
                for _ in range(3):
                    ws.cell(row=row, column=col, value="").alignment = center
                    col += 1
        
        # Итоговые вычисления для этого уровня (в итоговых колонках)
        total_level_count = sum(total_modal_values[crit.id].count(level) for crit in all_crit)
        ws.cell(row=row, column=col, value=total_level_count).alignment = center
        col += 1
        # Остальные уровни (пустые)
        for _ in range(2):
            ws.cell(row=row, column=col, value="").alignment = center
            col += 1
        # Условные (пустые)
        ws.cell(row=row, column=col, value="").alignment = center
        col += 1
        # Проценты
        total_possible = len(all_crit) * total_topics_all
        percentage = (total_level_count / total_possible * 100) if total_possible > 0 else 0
        ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
        col += 1
        # Остальные проценты (пустые)
        for _ in range(3):
            ws.cell(row=row, column=col, value="").alignment = center
            col += 1
        # Точные, допустимые, условные (пустые для этой строки)
        for _ in range(3):
            ws.cell(row=row, column=col, value="").alignment = center
            col += 1
        # Точные %, допустимые %, условные % (пустые для этой строки)
        for _ in range(3):
            ws.cell(row=row, column=col, value="").alignment = center
            col += 1
        
        row += 1
    
    # Строка "Условные" для всех классов
    ws.cell(row=row, column=1, value="Условные").font = bold
    col = 2
    for crit in all_crit:
        count = total_conditional_counts[crit.id]
        ws.cell(row=row, column=col, value=count)
        ws.cell(row=row, column=col).alignment = center
        col += 1
        
        # После последнего критерия в блоке добавляем вычисления для блока
        block_name = crit_to_block.get(crit.id, "Без блока")
        block_crits = [c for c in all_crit if crit_to_block.get(c.id) == block_name]
        if crit == block_crits[-1]:  # Последний критерий в блоке
            # Собираем условные по блоку
            block_conditional = sum(total_conditional_counts[c.id] for c in block_crits)
            num_crits_in_block = len(block_crits)
            # Уровни (пустые для этой строки)
            for _ in range(3):
                ws.cell(row=row, column=col, value="").alignment = center
                col += 1
            # Условные
            ws.cell(row=row, column=col, value=block_conditional).alignment = center
            col += 1
            # Проценты для уровней (пустые)
            for _ in range(3):
                ws.cell(row=row, column=col, value="").alignment = center
                col += 1
            # Процент для условных
            percentage = (block_conditional / (num_crits_in_block * total_topics_all) * 100) if (num_crits_in_block * total_topics_all) > 0 else 0
            ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
            col += 1
            # Точные, допустимые, условные (пустые для этой строки)
            for _ in range(3):
                ws.cell(row=row, column=col, value="").alignment = center
                col += 1
            # Точные %, допустимые %, условные % (пустые для этой строки)
            for _ in range(3):
                ws.cell(row=row, column=col, value="").alignment = center
                col += 1
    
    # Итоговые вычисления для условных
    total_conditional = sum(total_conditional_counts[crit.id] for crit in all_crit)
    # Уровни (пустые)
    for _ in range(3):
        ws.cell(row=row, column=col, value="").alignment = center
        col += 1
    # Условные
    ws.cell(row=row, column=col, value=total_conditional).alignment = center
    col += 1
    # Проценты для уровней (пустые)
    for _ in range(3):
        ws.cell(row=row, column=col, value="").alignment = center
        col += 1
    # Процент для условных
    total_possible = len(all_crit) * total_topics_all
    percentage = (total_conditional / total_possible * 100) if total_possible > 0 else 0
    ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
    col += 1
    # Точные, допустимые, условные (пустые для этой строки)
    for _ in range(3):
        ws.cell(row=row, column=col, value="").alignment = center
        col += 1
    # Точные %, допустимые %, условные % (пустые для этой строки)
    for _ in range(3):
        ws.cell(row=row, column=col, value="").alignment = center
        col += 1
    
    row += 1
    
    # Строки с процентами для всех классов
    if total_topics_all > 0:
        # Проценты для уровней
        for level in [1, 2, 3]:
            ws.cell(row=row, column=1, value=f"Уровень {level} %").font = bold
            col = 2
            for crit in all_crit:
                count = total_modal_values[crit.id].count(level)
                percentage = (count / total_topics_all * 100) if total_topics_all > 0 else 0
                ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%")
                ws.cell(row=row, column=col).alignment = center
                col += 1
                
                # После последнего критерия в блоке добавляем вычисления для блока
                block_name = crit_to_block.get(crit.id, "Без блока")
                block_crits = [c for c in all_crit if crit_to_block.get(c.id) == block_name]
                if crit == block_crits[-1]:  # Последний критерий в блоке
                    # Собираем данные по блоку для этого уровня
                    block_modal_values = []
                    for c in block_crits:
                        block_modal_values.extend([v for v in total_modal_values[c.id] if v == level])
                    block_count = len(block_modal_values)
                    num_crits_in_block = len(block_crits)
                    # Уровни 1-3 (проценты)
                    percentage = (block_count / (num_crits_in_block * total_topics_all) * 100) if (num_crits_in_block * total_topics_all) > 0 else 0
                    ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
                    col += 1
                    # Остальные уровни (пустые)
                    for _ in range(2):
                        ws.cell(row=row, column=col, value="").alignment = center
                        col += 1
                    # Условные (пустые)
                    ws.cell(row=row, column=col, value="").alignment = center
                    col += 1
                    # Проценты для уровней (пустые, так как это уже проценты)
                    for _ in range(3):
                        ws.cell(row=row, column=col, value="").alignment = center
                        col += 1
                    # Процент для условных (пустой)
                    ws.cell(row=row, column=col, value="").alignment = center
                    col += 1
                    # Точные, допустимые, условные (пустые для этой строки)
                    for _ in range(3):
                        ws.cell(row=row, column=col, value="").alignment = center
                        col += 1
                    # Точные %, допустимые %, условные % (пустые для этой строки)
                    for _ in range(3):
                        ws.cell(row=row, column=col, value="").alignment = center
                        col += 1
            
            # Итоговые вычисления для этого уровня
            total_level_count = sum(total_modal_values[crit.id].count(level) for crit in all_crit)
            total_possible = len(all_crit) * total_topics_all
            percentage = (total_level_count / total_possible * 100) if total_possible > 0 else 0
            ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
            col += 1
            # Остальные уровни (пустые)
            for _ in range(2):
                ws.cell(row=row, column=col, value="").alignment = center
                col += 1
            # Условные (пустые)
            ws.cell(row=row, column=col, value="").alignment = center
            col += 1
            # Проценты (пустые)
            for _ in range(4):
                ws.cell(row=row, column=col, value="").alignment = center
                col += 1
            # Точные, допустимые, условные (пустые для этой строки)
            for _ in range(3):
                ws.cell(row=row, column=col, value="").alignment = center
                col += 1
            # Точные %, допустимые %, условные % (пустые для этой строки)
            for _ in range(3):
                ws.cell(row=row, column=col, value="").alignment = center
                col += 1
            
            row += 1
        
        # Процент для условных
        ws.cell(row=row, column=1, value="Условные %").font = bold
        col = 2
        for crit in all_crit:
            count = total_conditional_counts[crit.id]
            percentage = (count / total_topics_all * 100) if total_topics_all > 0 else 0
            ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%")
            ws.cell(row=row, column=col).alignment = center
            col += 1
            
            # После последнего критерия в блоке добавляем вычисления для блока
            block_name = crit_to_block.get(crit.id, "Без блока")
            block_crits = [c for c in all_crit if crit_to_block.get(c.id) == block_name]
            if crit == block_crits[-1]:  # Последний критерий в блоке
                # Собираем условные по блоку
                block_conditional = sum(total_conditional_counts[c.id] for c in block_crits)
                num_crits_in_block = len(block_crits)
                # Уровни (пустые)
                for _ in range(3):
                    ws.cell(row=row, column=col, value="").alignment = center
                    col += 1
                # Условные (пустые, так как это уже проценты)
                ws.cell(row=row, column=col, value="").alignment = center
                col += 1
                # Проценты для уровней (пустые)
                for _ in range(3):
                    ws.cell(row=row, column=col, value="").alignment = center
                    col += 1
                # Процент для условных
                percentage = (block_conditional / (num_crits_in_block * total_topics_all) * 100) if (num_crits_in_block * total_topics_all) > 0 else 0
                ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
                col += 1
                # Точные, допустимые, условные (пустые для этой строки)
                for _ in range(3):
                    ws.cell(row=row, column=col, value="").alignment = center
                    col += 1
                # Точные %, допустимые %, условные % (пустые для этой строки)
                for _ in range(3):
                    ws.cell(row=row, column=col, value="").alignment = center
                    col += 1
        
        # Итоговые вычисления для условных
        total_conditional = sum(total_conditional_counts[crit.id] for crit in all_crit)
        # Уровни (пустые)
        for _ in range(3):
            ws.cell(row=row, column=col, value="").alignment = center
            col += 1
        # Условные (пустые)
        ws.cell(row=row, column=col, value="").alignment = center
        col += 1
        # Проценты для уровней (пустые)
        for _ in range(3):
            ws.cell(row=row, column=col, value="").alignment = center
            col += 1
        # Процент для условных
        total_possible = len(all_crit) * total_topics_all
        percentage = (total_conditional / total_possible * 100) if total_possible > 0 else 0
        ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
        
        row += 1
        
        # Подсчет точных, допустимых и условных значений для всех классов
        for degree_key, label, fill_color in [("exact", "Точные", green_fill), ("acceptable", "Допустимые", yellow_fill), ("conditional", "Условные", red_fill)]:
            ws.cell(row=row, column=1, value=label).font = bold
            ws.cell(row=row, column=1).fill = fill_color  # Закрашиваем ячейку названия
            col = 2
            for crit in all_crit:
                count = total_degree_counts[crit.id][degree_key]
                ws.cell(row=row, column=col, value=count)
                ws.cell(row=row, column=col).fill = fill_color
                ws.cell(row=row, column=col).alignment = center
                col += 1
                
                # После последнего критерия в блоке добавляем вычисления для блока
                block_name = crit_to_block.get(crit.id, "Без блока")
                block_crits = [c for c in all_crit if crit_to_block.get(c.id) == block_name]
                if crit == block_crits[-1]:  # Последний критерий в блоке
                    # Собираем данные по блоку для этой степени
                    block_count = sum(total_degree_counts[c.id][degree_key] for c in block_crits)
                    num_crits_in_block = len(block_crits)
                    # Уровни (пустые для этой строки)
                    for _ in range(3):
                        ws.cell(row=row, column=col, value="").alignment = center
                        col += 1
                    # Условные (пустые для этой строки, кроме строки "Условные")
                    if degree_key == "conditional":
                        ws.cell(row=row, column=col, value=block_count).fill = fill_color
                        ws.cell(row=row, column=col).alignment = center
                    else:
                        ws.cell(row=row, column=col, value="").alignment = center
                    col += 1
                    # Проценты для уровней (пустые)
                    for _ in range(3):
                        ws.cell(row=row, column=col, value="").alignment = center
                        col += 1
                    # Процент для условных (пустой, кроме строки "Условные")
                    if degree_key == "conditional":
                        percentage = (block_count / (num_crits_in_block * total_topics_all) * 100) if (num_crits_in_block * total_topics_all) > 0 else 0
                        ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
                    else:
                        ws.cell(row=row, column=col, value="").alignment = center
                    col += 1
                    # Точные, допустимые, условные (числа)
                    if degree_key == "exact":
                        ws.cell(row=row, column=col, value=block_count).fill = fill_color
                        ws.cell(row=row, column=col).alignment = center
                    elif degree_key == "acceptable":
                        ws.cell(row=row, column=col, value=block_count).fill = fill_color
                        ws.cell(row=row, column=col).alignment = center
                    elif degree_key == "conditional":
                        ws.cell(row=row, column=col, value=block_count).fill = fill_color
                        ws.cell(row=row, column=col).alignment = center
                    else:
                        ws.cell(row=row, column=col, value="").alignment = center
                    col += 1
                    # Остальные две колонки (пустые для этой строки)
                    for _ in range(2):
                        ws.cell(row=row, column=col, value="").alignment = center
                        col += 1
                    # Точные %, допустимые %, условные % (проценты)
                    if degree_key == "exact":
                        percentage = (block_count / (num_crits_in_block * total_topics_all) * 100) if (num_crits_in_block * total_topics_all) > 0 else 0
                        ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
                    elif degree_key == "acceptable":
                        percentage = (block_count / (num_crits_in_block * total_topics_all) * 100) if (num_crits_in_block * total_topics_all) > 0 else 0
                        ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
                    elif degree_key == "conditional":
                        # Процент уже выведен выше
                        ws.cell(row=row, column=col, value="").alignment = center
                    else:
                        ws.cell(row=row, column=col, value="").alignment = center
                    col += 1
                    # Остальные две колонки процентов (пустые для этой строки)
                    for _ in range(2):
                        ws.cell(row=row, column=col, value="").alignment = center
                        col += 1
            
            # Итоговые вычисления для этой степени
            total_count = sum(total_degree_counts[crit.id][degree_key] for crit in all_crit)
            # Уровни (пустые)
            for _ in range(3):
                ws.cell(row=row, column=col, value="").alignment = center
                col += 1
            # Условные (пустые, кроме строки "Условные")
            if degree_key == "conditional":
                ws.cell(row=row, column=col, value=total_count).fill = fill_color
                ws.cell(row=row, column=col).alignment = center
            else:
                ws.cell(row=row, column=col, value="").alignment = center
            col += 1
            # Проценты для уровней (пустые)
            for _ in range(3):
                ws.cell(row=row, column=col, value="").alignment = center
                col += 1
            # Процент для условных (пустой, кроме строки "Условные")
            if degree_key == "conditional":
                total_possible = len(all_crit) * total_topics_all
                percentage = (total_count / total_possible * 100) if total_possible > 0 else 0
                ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
            else:
                ws.cell(row=row, column=col, value="").alignment = center
            col += 1
            # Точные, допустимые, условные (числа)
            if degree_key == "exact":
                ws.cell(row=row, column=col, value=total_count).fill = fill_color
                ws.cell(row=row, column=col).alignment = center
            elif degree_key == "acceptable":
                ws.cell(row=row, column=col, value=total_count).fill = fill_color
                ws.cell(row=row, column=col).alignment = center
            elif degree_key == "conditional":
                ws.cell(row=row, column=col, value=total_count).fill = fill_color
                ws.cell(row=row, column=col).alignment = center
            else:
                ws.cell(row=row, column=col, value="").alignment = center
            col += 1
            # Остальные две колонки (пустые для этой строки)
            for _ in range(2):
                ws.cell(row=row, column=col, value="").alignment = center
                col += 1
            # Точные %, допустимые %, условные % (проценты)
            if degree_key == "exact":
                total_possible = len(all_crit) * total_topics_all
                percentage = (total_count / total_possible * 100) if total_possible > 0 else 0
                ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
            elif degree_key == "acceptable":
                total_possible = len(all_crit) * total_topics_all
                percentage = (total_count / total_possible * 100) if total_possible > 0 else 0
                ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
            elif degree_key == "conditional":
                # Процент уже выведен выше
                ws.cell(row=row, column=col, value="").alignment = center
            else:
                ws.cell(row=row, column=col, value="").alignment = center
            col += 1
            # Остальные две колонки процентов (пустые для этой строки)
            for _ in range(2):
                ws.cell(row=row, column=col, value="").alignment = center
                col += 1
            
            row += 1
        
        # Проценты для точных, допустимых и условных
        for degree_key, label in [("exact", "Точные %"), ("acceptable", "Допустимые %"), ("conditional", "Условные %")]:
            ws.cell(row=row, column=1, value=label).font = bold
            col = 2
            for crit in all_crit:
                count = total_degree_counts[crit.id][degree_key]
                percentage = (count / total_topics_all * 100) if total_topics_all > 0 else 0
                ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%")
                ws.cell(row=row, column=col).alignment = center
                col += 1
                
                # После последнего критерия в блоке добавляем вычисления для блока
                block_name = crit_to_block.get(crit.id, "Без блока")
                block_crits = [c for c in all_crit if crit_to_block.get(c.id) == block_name]
                if crit == block_crits[-1]:  # Последний критерий в блоке
                    # Собираем данные по блоку для этой степени
                    block_count = sum(total_degree_counts[c.id][degree_key] for c in block_crits)
                    num_crits_in_block = len(block_crits)
                    # Уровни (пустые)
                    for _ in range(3):
                        ws.cell(row=row, column=col, value="").alignment = center
                        col += 1
                    # Условные (пустые, так как это уже проценты)
                    ws.cell(row=row, column=col, value="").alignment = center
                    col += 1
                    # Проценты для уровней (пустые, так как это уже проценты)
                    for _ in range(3):
                        ws.cell(row=row, column=col, value="").alignment = center
                        col += 1
                    # Процент для условных
                    if degree_key == "conditional":
                        percentage = (block_count / (num_crits_in_block * total_topics_all) * 100) if (num_crits_in_block * total_topics_all) > 0 else 0
                        ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
                    else:
                        ws.cell(row=row, column=col, value="").alignment = center
                    col += 1
                    # Точные, допустимые, условные (пустые для этой строки, так как это уже проценты)
                    for _ in range(3):
                        ws.cell(row=row, column=col, value="").alignment = center
                        col += 1
                    # Точные %, допустимые %, условные % (проценты)
                    if degree_key == "exact":
                        percentage = (block_count / (num_crits_in_block * total_topics_all) * 100) if (num_crits_in_block * total_topics_all) > 0 else 0
                        ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
                    elif degree_key == "acceptable":
                        percentage = (block_count / (num_crits_in_block * total_topics_all) * 100) if (num_crits_in_block * total_topics_all) > 0 else 0
                        ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
                    elif degree_key == "conditional":
                        # Процент уже выведен выше
                        ws.cell(row=row, column=col, value="").alignment = center
                    else:
                        ws.cell(row=row, column=col, value="").alignment = center
                    col += 1
                    # Остальные две колонки процентов (пустые для этой строки)
                    for _ in range(2):
                        ws.cell(row=row, column=col, value="").alignment = center
                        col += 1
            
            # Итоговые вычисления для этой степени
            total_count = sum(total_degree_counts[crit.id][degree_key] for crit in all_crit)
            # Уровни (пустые)
            for _ in range(3):
                ws.cell(row=row, column=col, value="").alignment = center
                col += 1
            # Условные (пустые)
            ws.cell(row=row, column=col, value="").alignment = center
            col += 1
            # Проценты для уровней (пустые)
            for _ in range(3):
                ws.cell(row=row, column=col, value="").alignment = center
                col += 1
            # Процент для условных
            if degree_key == "conditional":
                total_possible = len(all_crit) * total_topics_all
                percentage = (total_count / total_possible * 100) if total_possible > 0 else 0
                ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
            else:
                ws.cell(row=row, column=col, value="").alignment = center
            col += 1
            # Точные, допустимые, условные (пустые для этой строки, так как это уже проценты)
            for _ in range(3):
                ws.cell(row=row, column=col, value="").alignment = center
                col += 1
            # Точные %, допустимые %, условные % (проценты)
            if degree_key == "exact":
                total_possible = len(all_crit) * total_topics_all
                percentage = (total_count / total_possible * 100) if total_possible > 0 else 0
                ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
            elif degree_key == "acceptable":
                total_possible = len(all_crit) * total_topics_all
                percentage = (total_count / total_possible * 100) if total_possible > 0 else 0
                ws.cell(row=row, column=col, value=f"{round(percentage, 1)}%").alignment = center
            elif degree_key == "conditional":
                # Процент уже выведен выше
                ws.cell(row=row, column=col, value="").alignment = center
            else:
                ws.cell(row=row, column=col, value="").alignment = center
            col += 1
            # Остальные две колонки процентов (пустые для этой строки)
            for _ in range(2):
                ws.cell(row=row, column=col, value="").alignment = center
                col += 1
            
            row += 1

    # Применяем рамки ко всем строкам (включая итоговые)
    for r in range(start_row, row):
        for c in range(1, total_cols + 1):
            ws.cell(row=r, column=c).border = thin_border

    return row + 2


def build_modal_sheet(ws, programs, main_criteria, extra_criteria):
    row = 1
    for program in programs:
        row = build_modal_block(ws, row, program, main_criteria, extra_criteria)
    # автоширина
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                try:
                    length = len(str(cell.value))
                except Exception:
                    length = 0
                max_length = max(max_length, length)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)


def build_program_block(ws, start_row, program, main_criteria, extra_criteria):
    """
    Создаёт блок сводной таблицы для одной образовательной программы.
    Только основные критерии (main_criteria), без дополнительных и комментариев.
    Возвращает номер строки, на которой закончился блок, и карту комментариев.
    """
    topics = list(program.topics.all().order_by("order", "id"))

    # соберём всех оценивателей
    from evaluations.models import TopicEvaluation
    evals = TopicEvaluation.objects.filter(
        program_evaluation__program=program
    ).select_related("program_evaluation__evaluator")

    evaluators = sorted({te.program_evaluation.evaluator.full_name for te in evals})

    # если нет оценивателей — поставим один "-"
    if not evaluators:
        evaluators = ["-"]

    # количество колонок на один критерий
    cols_per_criterion = len(evaluators)

    # стили
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    row = start_row

    # -------------------------------
    # 0. Строка с полными ФИО оценщиков (самый верх)
    # -------------------------------
    # Вычисляем общее количество колонок для основных критериев
    total_cols = 2 + len(main_criteria) * cols_per_criterion
    
    cell = ws.cell(row=row, column=1, value="Evaluator (Full Name)")
    cell.font = bold
    cell.alignment = center
    
    col = 2
    # Выводим ФИО каждого оценщика только 1 раз
    # Структура колонок: для каждого критерия есть cols_per_criterion колонок (по одной на каждого оценщика)
    # Для Ev1 (ev_idx=0) колонки идут: 2, 2+cols_per_criterion, 2+2*cols_per_criterion, ...
    # Для Ev2 (ev_idx=1) колонки идут: 3, 3+cols_per_criterion, 3+2*cols_per_criterion, ...
    # Объединяем ФИО на все колонки всех критериев для этого оценщика
    # Но так как openpyxl не поддерживает объединение с шагом, 
    # и объединение на все колонки первого критерия создаст пересечения между оценщиками,
    # объединяем на все колонки первого критерия последовательно для каждого оценщика
    # Ev1: колонки 2 до 2+cols_per_criterion-1 (все колонки первого критерия)
    # Ev2: колонки следующего блока и т.д.
    # Но это неправильно, потому что колонки первого критерия это 2, 3, 4 для всех оценщиков
    
    # Правильное решение: объединять на все колонки первого критерия для каждого оценщика отдельно
    # Но чтобы избежать пересечений, нужно объединять колонки последовательно:
    # Для первого критерия: Ev1 объединяет колонки 2-4, но это займет все колонки первого критерия
    # Поэтому просто выводим ФИО без объединения, чтобы избежать ошибок
    for ev_idx, ev_full in enumerate(evaluators):
        # Колонка для этого оценщика в первом критерии
        start_col = col + ev_idx
        
        # Устанавливаем значение и стили без объединения
        cell = ws.cell(row=row, column=start_col, value=ev_full)
        cell.alignment = center
        cell.font = Font(size=9)
    
    row += 1

    # -------------------------------
    # 1. Заголовок Program
    # -------------------------------
    cell1 = ws.cell(row=row, column=1, value="Program")
    cell1.font = bold
    cell2 = ws.cell(row=row, column=2, value=program.name)
    cell2.font = bold
    # объединить ячейки Program A → до конца таблицы
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=total_cols)

    row += 1

    # -------------------------------
    # 1.1 Заголовок блоков критериев (только основные)
    # -------------------------------
    # Группируем только основные критерии по блокам
    blocks = {}
    for c in main_criteria:
        block = c.block
        key = (block.order if block else 9999, block.name if block else "Без блока", block.id if block else None)
        blocks.setdefault(key, []).append(c)

    # сортировка блоков по order/name
    blocks_sorted = sorted(blocks.items(), key=lambda k: (k[0][0], k[0][1]))

    ws.cell(row=row + 1, column=1, value="Topic").font = bold  # зарезервируем ниже; сейчас готовим колонки

    col = 2
    for (_order, block_name, _bid), crit_list in blocks_sorted:
        span = len(crit_list) * cols_per_criterion
        ws.merge_cells(
            start_row=row + 1,
            start_column=col,
            end_row=row + 1,
            end_column=col + span - 1
        )
        ws.cell(row=row + 1, column=col, value=block_name).alignment = center
        ws.cell(row=row + 1, column=col, value=block_name).font = bold
        col += span

    row += 2

    # -------------------------------
    # 2. Заголовок критериев (верхний)
    # -------------------------------
    ws.cell(row=row, column=1, value="Topic").font = bold

    col = 2

    # следуем тому же порядку блоков (только основные критерии)
    all_crit = []
    for _, crit_list in blocks_sorted:
        all_crit.extend(sorted(crit_list, key=lambda c: (c.order, c.id)))

    for crit in all_crit:
        if crit == "comment":
            title = "Comment"
        else:
            title = crit.name

        cell = ws.cell(row=row, column=col, value=title)
        cell.alignment = center
        cell.font = bold

        ws.merge_cells(
            start_row=row,
            start_column=col,
            end_row=row,
            end_column=col + cols_per_criterion - 1
        )
        col += cols_per_criterion

    row += 1

    # -------------------------------
    # 3. Строка с фамилиями оценщиков
    # -------------------------------
    ws.cell(row=row, column=1, value="Evaluator (Last Name)").font = bold
    ws.cell(row=row, column=1).alignment = center

    col = 2
    for _crit in all_crit:
        for ev_full in evaluators:
            ev_lastname = get_lastname(ev_full)
            ws.cell(row=row, column=col, value=ev_lastname)
            ws.cell(row=row, column=col).alignment = center
            ws.cell(row=row, column=col).font = Font(size=10)
            col += 1

    row += 1

    # -------------------------------
    # 4. Строки тем
    # -------------------------------
    # Получим карту: {(topic_id, evaluator_name, criterion_id) → value}
    answer_map = {}
    comments_map = {}  # Отдельная карта для комментариев

    for te in evals.prefetch_related("answers__criterion"):
        evname = te.program_evaluation.evaluator.full_name
        for ans in te.answers.all():
            key = (te.topic_id, evname, ans.criterion_id)
            value = ans.value if ans.value is not None else ("Yes" if ans.yes_no else "No")
            answer_map[key] = value

        # комментарий сохраняем отдельно
        if te.comment:
            key = (te.topic_id, evname, "comment")
            comments_map[key] = te.comment

    # Группируем темы по class_level, чтобы визуально разделить
    grouped_topics = {}
    for t in topics:
        cls = t.class_level or "Без класса"
        grouped_topics.setdefault(cls, []).append(t)

    for cls_name in sorted(grouped_topics.keys()):
        # строка-заголовок класса
        ws.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=total_cols
        )
        ws.cell(row=row, column=1, value=f"Class: {cls_name}").font = bold
        ws.cell(row=row, column=1).alignment = center
        row += 1

        for topic in grouped_topics[cls_name]:
            ws.cell(row=row, column=1, value=topic.name)
            
            col = 2
            for crit in all_crit:
                for ev_full in evaluators:
                    key = (topic.id, ev_full, crit.id)
                    ws.cell(row=row, column=col, value=answer_map.get(key, "-"))
                    ws.cell(row=row, column=col).alignment = center
                    col += 1
            
            row += 1

    # рамки по всему блоку
    for r in range(start_row, row):
        for c in range(1, total_cols + 1):
            ws.cell(row=r, column=c).border = thin_border

    return row + 1, comments_map  # +1 чтобы оставить пустую строку, и карта комментариев


def build_all_programs_summary(ws, start_row, programs, main_criteria):
    """
    Строит сводную таблицу по всем программам:
    - Программы по вертикали
    - Критерии по горизонтали
    - Для каждой программы: количество оценок по уровням (1, 2, 3), их проценты
    - Итоговая строка за все программы
    
    Использует все оценки всех оценщиков (без закреплений и модальных значений).
    Для каждого критерия суммирует все оценки по каждому уровню.
    """
    from evaluations.models import TopicEvaluation
    
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Получаем все критерии
    all_crit = sorted(main_criteria, key=lambda c: (c.order if c.order else 9999, c.id))
    
    # Собираем данные для каждой программы
    program_data = {}  # {program_id: {crit_id: {'level_1': count, 'level_2': count, 'level_3': count, 'total': count}}}
    total_data = {}  # Итоговые данные по всем программам
    
    for crit in all_crit:
        total_data[crit.id] = {'level_1': 0, 'level_2': 0, 'level_3': 0, 'total': 0}
    
    for program in programs:
        program_data[program.id] = {}
        for crit in all_crit:
            program_data[program.id][crit.id] = {'level_1': 0, 'level_2': 0, 'level_3': 0, 'total': 0}
        
        # Получаем все оценки всех оценщиков для этой программы
        topic_evals = TopicEvaluation.objects.filter(
            program_evaluation__program=program,
            completed_at__isnull=False
        ).select_related("topic", "program_evaluation__evaluator").prefetch_related("answers__criterion")
        
        # Собираем все значения для каждого критерия
        for te in topic_evals:
            for ans in te.answers.all():
                crit_id = ans.criterion_id
                if crit_id not in [c.id for c in all_crit]:
                    continue
                
                # Получаем значение ответа
                val = ans.value if ans.value is not None else ("Yes" if ans.yes_no else "No")
                
                # Проверяем, является ли значение уровнем (1, 2, 3)
                try:
                    val_int = int(val)
                    if val_int in [1, 2, 3]:
                        program_data[program.id][crit_id][f'level_{val_int}'] += 1
                        program_data[program.id][crit_id]['total'] += 1
                        total_data[crit_id][f'level_{val_int}'] += 1
                        total_data[crit_id]['total'] += 1
                except (ValueError, TypeError):
                    # Если значение не является числом 1-3, не учитываем
                    pass
    
    # Строим таблицу
    row = start_row
    
    # Заголовок
    ws.cell(row=row, column=1, value="Программа").font = bold
    ws.cell(row=row, column=1).alignment = center
    col = 2
    
    # Заголовки критериев с колонками для уровней
    for crit in all_crit:
        # Объединяем ячейки для названия критерия
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 6)
        cell = ws.cell(row=row, column=col, value=crit.name)
        cell.font = bold
        cell.alignment = center
        col += 7
    
    row += 1
    
    # Подзаголовки: Ур.1, Ур.2, Ур.3, Ур.1%, Ур.2%, Ур.3%, Итого
    ws.cell(row=row, column=1, value="").font = bold
    col = 2
    for crit in all_crit:
        ws.cell(row=row, column=col, value="Ур.1").font = bold
        ws.cell(row=row, column=col).alignment = center
        col += 1
        ws.cell(row=row, column=col, value="Ур.2").font = bold
        ws.cell(row=row, column=col).alignment = center
        col += 1
        ws.cell(row=row, column=col, value="Ур.3").font = bold
        ws.cell(row=row, column=col).alignment = center
        col += 1
        ws.cell(row=row, column=col, value="Ур.1%").font = bold
        ws.cell(row=row, column=col).alignment = center
        col += 1
        ws.cell(row=row, column=col, value="Ур.2%").font = bold
        ws.cell(row=row, column=col).alignment = center
        col += 1
        ws.cell(row=row, column=col, value="Ур.3%").font = bold
        ws.cell(row=row, column=col).alignment = center
        col += 1
        ws.cell(row=row, column=col, value="Итого").font = bold
        ws.cell(row=row, column=col).alignment = center
        col += 1
    
    row += 1
    
    # Данные по программам
    for program in programs:
        ws.cell(row=row, column=1, value=program.name).font = bold
        col = 2
        
        for crit in all_crit:
            data = program_data[program.id][crit.id]
            total = data['total']
            
            # Количество по уровням
            ws.cell(row=row, column=col, value=data['level_1']).alignment = center
            col += 1
            ws.cell(row=row, column=col, value=data['level_2']).alignment = center
            col += 1
            ws.cell(row=row, column=col, value=data['level_3']).alignment = center
            col += 1
            
            # Проценты по уровням
            pct1 = (data['level_1'] / total * 100) if total > 0 else 0
            pct2 = (data['level_2'] / total * 100) if total > 0 else 0
            pct3 = (data['level_3'] / total * 100) if total > 0 else 0
            ws.cell(row=row, column=col, value=f"{round(pct1, 1)}%").alignment = center
            col += 1
            ws.cell(row=row, column=col, value=f"{round(pct2, 1)}%").alignment = center
            col += 1
            ws.cell(row=row, column=col, value=f"{round(pct3, 1)}%").alignment = center
            col += 1
            
            # Итого
            ws.cell(row=row, column=col, value=total).alignment = center
            col += 1
        
        row += 1
    
    # Итоговая строка
    ws.cell(row=row, column=1, value="Итого").font = bold
    col = 2
    
    for crit in all_crit:
        data = total_data[crit.id]
        total = data['total']
        
        # Количество по уровням
        ws.cell(row=row, column=col, value=data['level_1']).font = bold
        ws.cell(row=row, column=col).alignment = center
        col += 1
        ws.cell(row=row, column=col, value=data['level_2']).font = bold
        ws.cell(row=row, column=col).alignment = center
        col += 1
        ws.cell(row=row, column=col, value=data['level_3']).font = bold
        ws.cell(row=row, column=col).alignment = center
        col += 1
        
        # Проценты по уровням
        pct1 = (data['level_1'] / total * 100) if total > 0 else 0
        pct2 = (data['level_2'] / total * 100) if total > 0 else 0
        pct3 = (data['level_3'] / total * 100) if total > 0 else 0
        ws.cell(row=row, column=col, value=f"{round(pct1, 1)}%").font = bold
        ws.cell(row=row, column=col).alignment = center
        col += 1
        ws.cell(row=row, column=col, value=f"{round(pct2, 1)}%").font = bold
        ws.cell(row=row, column=col).alignment = center
        col += 1
        ws.cell(row=row, column=col, value=f"{round(pct3, 1)}%").font = bold
        ws.cell(row=row, column=col).alignment = center
        col += 1
        
        # Итого
        ws.cell(row=row, column=col, value=total).font = bold
        ws.cell(row=row, column=col).alignment = center
        col += 1
    
    # Рамки
    total_cols = 1 + len(all_crit) * 7
    for r in range(start_row, row + 1):
        for c in range(1, total_cols + 1):
            ws.cell(row=r, column=c).border = thin_border
    
    return row + 1


def build_comments_sheet(ws, program, comments_map, evaluators):
    """Создаёт лист с комментариями для программы."""
    from openpyxl.styles import Font, Alignment, Border, Side
    
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    row = 1
    
    # Заголовок
    ws.cell(row=row, column=1, value="Program").font = bold
    cell = ws.cell(row=row, column=2, value=program.name)
    cell.font = bold
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    row += 2
    
    # Заголовки колонок
    ws.cell(row=row, column=1, value="Topic").font = bold
    ws.cell(row=row, column=2, value="Evaluator").font = bold
    ws.cell(row=row, column=3, value="Comment").font = bold
    row += 1
    
    # Группируем темы по классам
    topics = list(program.topics.all().order_by("order", "id"))
    grouped_topics = {}
    for t in topics:
        cls = t.class_level or "Без класса"
        grouped_topics.setdefault(cls, []).append(t)
    
    for cls_name in sorted(grouped_topics.keys()):
        # Заголовок класса
        cell = ws.cell(row=row, column=1, value=f"Class: {cls_name}")
        cell.font = bold
        cell.alignment = center
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1
        
        for topic in grouped_topics[cls_name]:
            for ev_full in evaluators:
                key = (topic.id, ev_full, "comment")
                comment = comments_map.get(key)
                if comment:
                    ws.cell(row=row, column=1, value=topic.name)
                    ws.cell(row=row, column=2, value=ev_full)
                    ws.cell(row=row, column=3, value=comment)
                    row += 1
    
    # Рамки
    for r in range(1, row):
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = thin_border
    
    # Автоширина
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 60


def build_extra_criteria_block(ws, start_row, program, extra_criteria):
    """
    Создаёт блок для дополнительных критериев (extra_criteria).
    Структура похожа на build_program_block, но только для дополнительных критериев.
    """
    topics = list(program.topics.all().order_by("order", "id"))

    # соберём всех оценивателей
    from evaluations.models import TopicEvaluation
    evals = TopicEvaluation.objects.filter(
        program_evaluation__program=program
    ).select_related("program_evaluation__evaluator")

    evaluators = sorted({te.program_evaluation.evaluator.full_name for te in evals})

    # если нет оценивателей — поставим один "-"
    if not evaluators:
        evaluators = ["-"]

    # количество колонок на один критерий
    cols_per_criterion = len(evaluators)

    # стили
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    row = start_row

    # -------------------------------
    # 0. Строка с полными ФИО оценщиков (самый верх)
    # -------------------------------
    total_cols = 2 + len(extra_criteria) * cols_per_criterion
    
    ws.cell(row=row, column=1, value="Evaluator (Full Name)").font = bold
    ws.cell(row=row, column=1).alignment = center
    
    col = 2
    for ev_full in evaluators:
        cell = ws.cell(row=row, column=col, value=ev_full)
        cell.alignment = center
        cell.font = Font(size=9)
        ws.merge_cells(
            start_row=row,
            start_column=col,
            end_row=row,
            end_column=col + 2
        )
        col += 3
    
    row += 1

    # -------------------------------
    # 1. Заголовок Program
    # -------------------------------
    cell1 = ws.cell(row=row, column=1, value="Program")
    cell1.font = bold
    cell2 = ws.cell(row=row, column=2, value=program.name)
    cell2.font = bold
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=total_cols)
    row += 1

    # -------------------------------
    # 1.1 Заголовок блоков критериев (только дополнительные)
    # -------------------------------
    blocks = {}
    for c in extra_criteria:
        block = c.block
        key = (block.order if block else 9999, block.name if block else "Без блока", block.id if block else None)
        blocks.setdefault(key, []).append(c)

    blocks_sorted = sorted(blocks.items(), key=lambda k: (k[0][0], k[0][1]))

    ws.cell(row=row + 1, column=1, value="Topic").font = bold

    col = 2
    for (_order, block_name, _bid), crit_list in blocks_sorted:
        span = len(crit_list) * cols_per_criterion
        cell = ws.cell(row=row + 1, column=col, value=block_name)
        cell.alignment = center
        cell.font = bold
        ws.merge_cells(
            start_row=row + 1,
            start_column=col,
            end_row=row + 1,
            end_column=col + span - 1
        )
        col += span

    row += 2

    # -------------------------------
    # 2. Заголовок критериев
    # -------------------------------
    ws.cell(row=row, column=1, value="Topic").font = bold
    col = 2
    all_crit = []
    for _, crit_list in blocks_sorted:
        all_crit.extend(sorted(crit_list, key=lambda c: (c.order, c.id)))

    for crit in all_crit:
        cell = ws.cell(row=row, column=col, value=crit.name)
        cell.alignment = center
        cell.font = bold
        ws.merge_cells(
            start_row=row,
            start_column=col,
            end_row=row,
            end_column=col + cols_per_criterion - 1
        )
        col += cols_per_criterion

    row += 1

    # -------------------------------
    # 3. Строка с фамилиями оценщиков
    # -------------------------------
    ws.cell(row=row, column=1, value="Evaluator (Last Name)").font = bold
    ws.cell(row=row, column=1).alignment = center
    col = 2
    for _crit in all_crit:
        for ev_full in evaluators:
            ev_lastname = get_lastname(ev_full)
            ws.cell(row=row, column=col, value=ev_lastname)
            ws.cell(row=row, column=col).alignment = center
            ws.cell(row=row, column=col).font = Font(size=10)
            col += 1
    row += 1

    # -------------------------------
    # 4. Строки тем
    # -------------------------------
    from evaluations.models import Criterion
    answer_map = {}
    for te in evals.prefetch_related("answers__criterion"):
        evname = te.program_evaluation.evaluator.full_name
        for ans in te.answers.all():
            if ans.criterion.type == Criterion.EXTRA:
                key = (te.topic_id, evname, ans.criterion_id)
                value = ans.value if ans.value is not None else ("Yes" if ans.yes_no else "No")
                answer_map[key] = value

    grouped_topics = {}
    for t in topics:
        cls = t.class_level or "Без класса"
        grouped_topics.setdefault(cls, []).append(t)

    for cls_name in sorted(grouped_topics.keys()):
        cell = ws.cell(row=row, column=1, value=f"Class: {cls_name}")
        cell.font = bold
        cell.alignment = center
        ws.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=total_cols
        )
        row += 1

        for topic in grouped_topics[cls_name]:
            ws.cell(row=row, column=1, value=topic.name)
            col = 2
            for crit in all_crit:
                for ev_full in evaluators:
                    key = (topic.id, ev_full, crit.id)
                    ws.cell(row=row, column=col, value=answer_map.get(key, "-"))
                    ws.cell(row=row, column=col).alignment = center
                    col += 1
            row += 1

    # рамки
    for r in range(start_row, row):
        for c in range(1, total_cols + 1):
            ws.cell(row=r, column=c).border = thin_border

    return row + 1


@login_required
def export_program_xlsx(request, program_id):
    if not is_subadmin(request.user):
        raise PermissionDenied

    program = Program.objects.get(id=program_id)

    wb = Workbook()
    ws = wb.active
    ws.title = program.name[:31]

    from evaluations.models import Criterion, TopicEvaluation
    main_criteria = list(Criterion.objects.filter(type=Criterion.MAIN))
    extra_criteria = list(Criterion.objects.filter(type=Criterion.EXTRA))

    # Собираем всех оценщиков
    evals = TopicEvaluation.objects.filter(
        program_evaluation__program=program
    ).select_related("program_evaluation__evaluator")
    evaluators = sorted({te.program_evaluation.evaluator.full_name for te in evals})

    # Строим основной лист (только основные критерии)
    _, comments_map = build_program_block(ws, 1, program, main_criteria, extra_criteria)

    # --- Лист с комментариями ---
    comments_ws = wb.create_sheet("Comments")
    build_comments_sheet(comments_ws, program, comments_map, evaluators)

    # --- Лист с дополнительными критериями ---
    if extra_criteria:
        extra_ws = wb.create_sheet("Extra Criteria")
        build_extra_criteria_block(extra_ws, 1, program, extra_criteria)
        
        # Автоширина для листа с дополнительными критериями
        for col_idx in range(1, extra_ws.max_column + 1):
            max_length = 0
            for row_idx in range(1, extra_ws.max_row + 1):
                cell = extra_ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    try:
                        length = len(str(cell.value))
                    except Exception:
                        length = 0
                    if length > max_length:
                        max_length = length
            col_letter = get_column_letter(col_idx)
            extra_ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    # автоширина
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                # cast to str and measure
                try:
                    length = len(str(cell.value))
                except Exception:
                    length = 0
                if length > max_length:
                    max_length = length
        col_letter = get_column_letter(col_idx)
        # ограничим ширину для читаемости (например, 50)
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    response = HttpResponse(
        bio.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename=program_{program.id}_summary.xlsx'

    log_action(request.user, 'export_program_xlsx', object_type='Program', object_id=program.id)
    return response


@login_required
def export_modal_xlsx(request, program_id):
    """Отдельный экспорт модальных значений для программы."""
    if not is_subadmin(request.user):
        raise PermissionDenied

    from django.contrib import messages
    from evaluations.models import EvaluatorAssignment
    
    program = Program.objects.get(id=program_id)
    
    # Проверяем, что для всех классов есть закрепления
    topics = program.topics.all()
    class_levels = set(t.class_level or "Без класса" for t in topics)
    
    assignments = EvaluatorAssignment.objects.filter(program=program)
    assigned_classes = set(a.class_level for a in assignments)
    
    missing_classes = class_levels - assigned_classes
    
    if missing_classes:
        messages.error(
            request,
            f'Необходимо закрепить оценщиков для всех классов программы. '
            f'Отсутствуют закрепления для классов: {", ".join(sorted(missing_classes))}. '
            f'Пожалуйста, перейдите на страницу "Закрепление оценщиков".'
        )
        return redirect('program_detail', pk=program_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Modal"

    from evaluations.models import Criterion
    main_criteria = list(Criterion.objects.filter(type=Criterion.MAIN))
    extra_criteria = list(Criterion.objects.filter(type=Criterion.EXTRA))

    build_modal_sheet(ws, [program], main_criteria, extra_criteria)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    response = HttpResponse(
        bio.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename=program_{program.id}_modal.xlsx'

    log_action(request.user, 'export_modal_xlsx', object_type='Program', object_id=program.id)
    return response


@login_required
def assign_evaluators(request, pk):
    """Выбор и сохранение 3 оценщиков для каждого класса программы."""
    if not is_subadmin(request.user):
        raise PermissionDenied
    
    from evaluations.models import EvaluatorAssignment, EvaluatorSession, TopicEvaluation
    
    program = get_object_or_404(Program, pk=pk)
    
    # Получаем все классы тем в программе
    topics = program.topics.all()
    class_levels = sorted(set(t.class_level or "Без класса" for t in topics))
    
    # Получаем всех оценщиков, которые оценили хотя бы одну тему в программе
    topic_evals = TopicEvaluation.objects.filter(
        program_evaluation__program=program,
        completed_at__isnull=False
    ).select_related('program_evaluation__evaluator')
    evaluators = sorted(
        {te.program_evaluation.evaluator for te in topic_evals},
        key=lambda e: e.full_name
    )
    
    # Получаем существующие закрепления
    assignments = EvaluatorAssignment.objects.filter(program=program)
    existing_assignments = {a.class_level: a for a in assignments}
    
    if request.method == 'POST':
        # Сохраняем закрепления
        for class_level in class_levels:
            eval1_id = request.POST.get(f'evaluator1_{class_level}')
            eval2_id = request.POST.get(f'evaluator2_{class_level}')
            eval3_id = request.POST.get(f'evaluator3_{class_level}')
            
            if eval1_id and eval2_id and eval3_id:
                try:
                    eval1 = EvaluatorSession.objects.get(id=eval1_id)
                    eval2 = EvaluatorSession.objects.get(id=eval2_id)
                    eval3 = EvaluatorSession.objects.get(id=eval3_id)
                    
                    # Проверяем, что все разные
                    if len({eval1, eval2, eval3}) != 3:
                        continue
                    
                    assignment, created = EvaluatorAssignment.objects.update_or_create(
                        program=program,
                        class_level=class_level,
                        defaults={
                            'evaluator1': eval1,
                            'evaluator2': eval2,
                            'evaluator3': eval3,
                        }
                    )
                    
                    log_action(
                        request.user,
                        'assign_evaluators' if created else 'update_evaluators',
                        'EvaluatorAssignment',
                        assignment.id,
                        description=f'Закрепление оценщиков для класса {class_level} в программе {program.name}'
                    )
                except EvaluatorSession.DoesNotExist:
                    continue
        
        return redirect('program_detail', pk=pk)
    
    # Подготавливаем данные для шаблона
    assignments_data = []
    for class_level in class_levels:
        assignment = existing_assignments.get(class_level)
        assignments_data.append({
            'class_level': class_level,
            'evaluator1_id': assignment.evaluator1.id if assignment else None,
            'evaluator2_id': assignment.evaluator2.id if assignment else None,
            'evaluator3_id': assignment.evaluator3.id if assignment else None,
        })
    
    return render(request, 'manager/assign_evaluators.html', {
        'program': program,
        'assignments_data': assignments_data,
        'evaluators': evaluators,
        'breadcrumbs': [
            ('Программы', '/manager/'),
            (program.name, f'/manager/program/{program.id}/'),
            ('Закрепление оценщиков', '')
        ]
    })


@login_required
def export_program_csv(request, program_id):
    """Экспорт результатов программы в CSV для менеджера"""
    if not is_subadmin(request.user):
        raise PermissionDenied

    program = get_object_or_404(Program, id=program_id)

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="program_{program_id}_results.csv"'
    
    # Добавляем BOM для корректного отображения кириллицы в Excel
    response.write('\ufeff')
    
    writer = csv.writer(response)
    
    # Заголовки
    writer.writerow([
        "Программа",
        "Тема",
        "Оценивающий",
        "Критерий",
        "Оценка",
        "Да/Нет",
        "Комментарий",
        "Дата оценки",
    ])

    topic_evals = TopicEvaluation.objects.filter(
        program_evaluation__program=program,
        completed_at__isnull=False
    ).select_related('program_evaluation__evaluator', 'topic').prefetch_related('answers__criterion').order_by('topic__order', 'topic__id')

    for te in topic_evals:
        for ans in te.answers.all():
            value_str = str(ans.value) if ans.value is not None else ""
            # yes_no оставляем для обратной совместимости, но для extra критериев теперь используется value
            yes_no_str = "Да" if ans.yes_no else "Нет" if ans.yes_no is False else ""
            
            writer.writerow([
                program.name,
                te.topic.name,
                te.program_evaluation.evaluator.full_name,
                ans.criterion.name,
                value_str,
                yes_no_str,
                te.comment,
                te.completed_at.strftime('%Y-%m-%d %H:%M:%S') if te.completed_at else "",
            ])

    log_action(request.user, 'export_program_csv', object_type='Program', object_id=program.id)
    return response


@login_required
def export_all_xlsx(request):
    if not is_subadmin(request.user):
        raise PermissionDenied

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    from evaluations.models import Criterion, Program
    main_criteria = list(Criterion.objects.filter(type=Criterion.MAIN))
    extra_criteria = list(Criterion.objects.filter(type=Criterion.EXTRA))

    programs = list(Program.objects.all().order_by("id"))
    
    # Строим сводную таблицу по всем программам
    build_all_programs_summary(ws, 1, programs, main_criteria)

    # автоширина колонок
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                # cast to str and measure
                try:
                    length = len(str(cell.value))
                except Exception:
                    length = 0
                if length > max_length:
                    max_length = length
        col_letter = get_column_letter(col_idx)
        # ограничим ширину для читаемости (например, 50)
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    # отдаём файл
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    response = HttpResponse(
        bio.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=all_programs_summary.xlsx'

    log_action(request.user, 'export_all_xlsx', object_type='Summary')
    return response


@login_required
def export_evaluators_xlsx(request):
    """Экспорт списка оценщиков по программам с классами тем, которые они оценили."""
    if not is_subadmin(request.user):
        raise PermissionDenied

    from evaluations.models import TopicEvaluation
    from collections import defaultdict
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Оценщики"

    # Стили
    bold = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    program_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    program_font = Font(bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    row = 1

    # Заголовок
    cell = ws.cell(row=row, column=1, value="Экспорт оценщиков")
    cell.font = Font(bold=True, size=14)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 2

    # Заголовки колонок
    headers = ["Программа", "Количество тем", "ФИО оценщика", "Номер телефона", "Оцененные классы тем"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
    row += 1

    # Собираем данные по программам
    programs = Program.objects.all().order_by("id")
    
    for program in programs:
        # Получаем общее количество тем в программе
        total_topics = program.topics.count()
        
        # Получаем всех оценщиков программы, которые оценили хотя бы одну тему
        topic_evals = TopicEvaluation.objects.filter(
            program_evaluation__program=program,
            completed_at__isnull=False
        ).select_related('program_evaluation__evaluator', 'topic')
        
        # Группируем по оценщикам и собираем классы тем и количество оцененных тем
        evaluator_classes = defaultdict(set)
        evaluator_topic_count = defaultdict(int)
        for te in topic_evals:
            evaluator = te.program_evaluation.evaluator
            class_level = te.topic.class_level or "Без класса"
            evaluator_classes[evaluator].add(class_level)
            evaluator_topic_count[evaluator] += 1
        
        # Если есть оценщики в программе
        if evaluator_classes:
            # Строка с названием программы
            program_cell = ws.cell(row=row, column=1, value=program.name)
            program_cell.font = program_font
            program_cell.fill = program_fill
            program_cell.alignment = center
            
            # Количество тем в программе
            topics_count_cell = ws.cell(row=row, column=2, value=total_topics)
            topics_count_cell.font = program_font
            topics_count_cell.fill = program_fill
            topics_count_cell.alignment = center
            
            # Объединяем остальные колонки
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = thin_border
            row += 1
            
            # Строки с оценщиками
            for evaluator in sorted(evaluator_classes.keys(), key=lambda e: e.full_name):
                classes_list = sorted(evaluator_classes[evaluator])
                classes_str = ", ".join(classes_list)
                evaluated_count = evaluator_topic_count[evaluator]
                
                ws.cell(row=row, column=1, value="")  # Пустая колонка программы
                ws.cell(row=row, column=2, value=evaluated_count)  # Количество оцененных тем
                ws.cell(row=row, column=3, value=evaluator.full_name)
                ws.cell(row=row, column=4, value=evaluator.phone)
                ws.cell(row=row, column=5, value=classes_str)
                
                # Применяем стили
                for col in range(1, 6):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    if col == 1:
                        cell.fill = program_fill
                    elif col == 2:
                        cell.alignment = center
                
                row += 1
            
            # Пустая строка между программами
            row += 1
    
    # Автоширина колонок
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 50

    # Высота строк заголовков
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[3].height = 20

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    response = HttpResponse(
        bio.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=evaluators_export.xlsx'

    log_action(request.user, 'export_evaluators_xlsx', object_type='Summary')
    return response


# Запрос выгрузки оценок программ (соответствует выгрузка оценок программ.txt)
SQL_EXPORT_EVALUATIONS = """
SELECT
    p.name       AS program_name,
    t.name       AS topic_name,
    t.class_level AS class_level,
    es.full_name AS evaluator_name,
    MAX(te.completed_at) AS "Время оценки",
    MAX(CASE WHEN c.id = 1  THEN a.value END) AS "Независимость и патриотизм",
    MAX(CASE WHEN c.id = 2  THEN a.value END) AS "Единство и солидарность",
    MAX(CASE WHEN c.id = 3  THEN a.value END) AS "Справедливость и ответственность",
    MAX(CASE WHEN c.id = 4  THEN a.value END) AS "Закон и порядок",
    MAX(CASE WHEN c.id = 5  THEN a.value END) AS "Трудолюбие и профессионализм",
    MAX(CASE WHEN c.id = 6  THEN a.value END) AS "Созидание и новаторство",
    MAX(CASE WHEN c.id = 7  THEN a.value END) AS "Принятие решений и прикладные задачи",
    MAX(CASE WHEN c.id = 8  THEN a.value END) AS "Креативное и критическое мышление",
    MAX(CASE WHEN c.id = 9  THEN a.value END) AS "Коммуникация и взаимодействие",
    MAX(CASE WHEN c.id = 10 THEN a.value END) AS "Самосознание и эмпатия",
    MAX(CASE WHEN c.id = 11 THEN a.value END) AS "Управление эмоциями и стрессом",
    MAX(CASE WHEN c.id = 12 THEN a.value END) AS "Использование ИИ",
    MAX(CASE WHEN c.id = 13 THEN a.value END) AS "Современные инфоресурсы",
    MAX(CASE WHEN c.id = 14 THEN a.value END) AS "Современные пед. подходы",
    MAX(CASE WHEN c.id = 16 THEN a.value END) AS "Автономность",
    MAX(CASE WHEN c.id = 17 THEN a.value END) AS "Компетентность",
    MAX(CASE WHEN c.id = 18 THEN a.value END) AS "Связанность и причастность",
    MAX(CASE WHEN c.id = 19 THEN a.value END) AS "Минимальные часы",
    MAX(CASE WHEN c.id = 20 THEN a.value END) AS "Оптимальные часы"
FROM evaluations_answer a
JOIN evaluations_criterion c ON c.id = a.criterion_id
JOIN evaluations_topicevaluation te ON te.id = a.topic_evaluation_id
JOIN evaluations_topic t ON t.id = te.topic_id
JOIN evaluations_programevaluation pe ON pe.id = te.program_evaluation_id
JOIN evaluations_program p ON p.id = pe.program_id
JOIN evaluations_evaluatorsession es ON es.id = pe.evaluator_id
GROUP BY p.name, t.name, t.class_level, es.full_name
ORDER BY p.name, es.full_name, MIN(t.%(order_col)s)
"""

# Медианное время оценки: MySQL
SQL_MEDIAN_TIME_MYSQL = """
WITH ordered_topics AS (
    SELECT
        es.full_name,
        p.name AS program_name,
        pe.id AS program_evaluation_id,
        te.completed_at,
        LAG(te.completed_at) OVER (PARTITION BY pe.id ORDER BY te.completed_at) AS prev_completed_at
    FROM evaluations_topicevaluation te
    JOIN evaluations_programevaluation pe ON te.program_evaluation_id = pe.id
    JOIN evaluations_program p ON pe.program_id = p.id
    JOIN evaluations_evaluatorsession es ON pe.evaluator_id = es.id
    WHERE te.completed_at IS NOT NULL
),
time_diffs AS (
    SELECT full_name, program_name,
        TIMESTAMPDIFF(SECOND, prev_completed_at, completed_at) AS diff_seconds
    FROM ordered_topics
    WHERE prev_completed_at IS NOT NULL
),
ranked AS (
    SELECT full_name, program_name, diff_seconds,
        ROW_NUMBER() OVER (PARTITION BY full_name, program_name ORDER BY diff_seconds) AS rn,
        COUNT(*) OVER (PARTITION BY full_name, program_name) AS total_count
    FROM time_diffs
)
SELECT full_name, program_name, AVG(diff_seconds) AS median_seconds
FROM ranked
WHERE rn IN (FLOOR((total_count + 1) / 2), FLOOR((total_count + 2) / 2))
GROUP BY full_name, program_name
ORDER BY full_name, program_name
"""

# Медианное время оценки: SQLite (julianday для разницы в секундах)
SQL_MEDIAN_TIME_SQLITE = """
WITH ordered_topics AS (
    SELECT
        es.full_name,
        p.name AS program_name,
        pe.id AS program_evaluation_id,
        te.completed_at,
        LAG(te.completed_at) OVER (PARTITION BY pe.id ORDER BY te.completed_at) AS prev_completed_at
    FROM evaluations_topicevaluation te
    JOIN evaluations_programevaluation pe ON te.program_evaluation_id = pe.id
    JOIN evaluations_program p ON pe.program_id = p.id
    JOIN evaluations_evaluatorsession es ON pe.evaluator_id = es.id
    WHERE te.completed_at IS NOT NULL
),
time_diffs AS (
    SELECT full_name, program_name,
        (julianday(completed_at) - julianday(prev_completed_at)) * 86400 AS diff_seconds
    FROM ordered_topics
    WHERE prev_completed_at IS NOT NULL
),
ranked AS (
    SELECT full_name, program_name, diff_seconds,
        ROW_NUMBER() OVER (PARTITION BY full_name, program_name ORDER BY diff_seconds) AS rn,
        COUNT(*) OVER (PARTITION BY full_name, program_name) AS total_count
    FROM time_diffs
)
SELECT full_name, program_name, AVG(diff_seconds) AS median_seconds
FROM ranked
WHERE rn IN (FLOOR((total_count + 1) / 2.0), FLOOR((total_count + 2) / 2.0))
GROUP BY full_name, program_name
ORDER BY full_name, program_name
"""


def _set_long_query_timeout(cursor):
    """Увеличить таймаут тяжёлого запроса (MySQL: 10 мин). При таймауте смотреть логи сервера и БД."""
    if connection.vendor == 'mysql':
        try:
            # MySQL 8.0.2+: max_execution_time в миллисекундах (600000 = 10 мин)
            cursor.execute("SET SESSION max_execution_time = 600000")
        except Exception:
            pass  # старые версии MySQL не поддерживают max_execution_time

@login_required
def export_evaluations_xlsx(request):
    """Экспорт оценок программ (выгрузка оценок программ) в Excel.
    При падении на сервере: проверить таймаут (БД, gunicorn/uwsgi, nginx) и логи."""
    if not is_subadmin(request.user):
        raise PermissionDenied
    order_col = '"order"' if connection.vendor == 'sqlite3' else '`order`'
    query = SQL_EXPORT_EVALUATIONS % {'order_col': order_col}
    with connection.cursor() as cursor:
        _set_long_query_timeout(cursor)
        cursor.execute(query)
        columns = [col[0] for col in cursor.description]
        rows = cursor.fetchall()
    wb = Workbook()
    ws = wb.active
    ws.title = "Оценки"
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for col_idx, header in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    def _add_tz_hours(val):
        """Добавить +5 ч к времени оценки (в БД хранится −5 ч). Работает с SQLite и MySQL."""
        if val is None:
            return val
        try:
            if hasattr(val, 'timestamp'):
                return val + timedelta(hours=5)
            s = str(val).strip()
            if not s:
                return val
            dt = parse_datetime(s)
            return (dt + timedelta(hours=5)) if dt is not None else val
        except (TypeError, ValueError):
            return val

    for row_idx, row_data in enumerate(rows, start=2):
        row_list = list(row_data)
        if len(row_list) >= 5:
            row_list[4] = _add_tz_hours(row_list[4])
        for col_idx, value in enumerate(row_list, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
    for col_idx in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    response = HttpResponse(
        bio.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=export_evaluations.xlsx'
    log_action(request.user, 'export_evaluations_xlsx', object_type='Summary')
    return response


@login_required
def export_evaluation_time_xlsx(request):
    """Экспорт медианного времени оценки по программам в Excel."""
    if not is_subadmin(request.user):
        raise PermissionDenied
    sql = SQL_MEDIAN_TIME_SQLITE if connection.vendor == 'sqlite3' else SQL_MEDIAN_TIME_MYSQL
    with connection.cursor() as cursor:
        _set_long_query_timeout(cursor)
        cursor.execute(sql)
        columns = [col[0] for col in cursor.description]
        rows = cursor.fetchall()
    wb = Workbook()
    ws = wb.active
    ws.title = "Время оценки"
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for col_idx, header in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 16
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    response = HttpResponse(
        bio.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=export_evaluation_time.xlsx'
    log_action(request.user, 'export_evaluation_time_xlsx', object_type='Summary')
    return response


@login_required
def manager_index(request):
    if not is_subadmin(request.user): raise PermissionDenied

    q = request.GET.get('q', '').strip()
    sort = request.GET.get('sort', 'name')

    programs = Program.objects.all()

    if q:
        programs = programs.filter(name__icontains=q)

    if sort:
        programs = programs.order_by(sort)

    # Добавляем статистику для каждой программы
    program_stats = []
    for program in programs:
        total_topics = program.topics.count()
        # Количество уникальных оценивших (тех, кто заполнил хотя бы одну тему)
        evaluators_count = TopicEvaluation.objects.filter(
            program_evaluation__program=program,
            completed_at__isnull=False
        ).values('program_evaluation__evaluator').distinct().count()
        # Количество заполненных оценок тем
        completed_evaluations = TopicEvaluation.objects.filter(
            program_evaluation__program=program,
            completed_at__isnull=False
        ).count()
        # Процент заполнения
        if total_topics > 0:
            completion_percent = round((completed_evaluations / (total_topics * max(evaluators_count, 1))) * 100, 1)
        else:
            completion_percent = 0
        
        program_stats.append({
            'program': program,
            'total_topics': total_topics,
            'evaluators_count': evaluators_count,
            'completed_evaluations': completed_evaluations,
            'completion_percent': completion_percent,
        })

    paginator = Paginator(program_stats, 20)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)

    return render(request, 'manager/index.html', {
        'page_obj': page_obj,
        'q': q,
        'sort': sort,
        'breadcrumbs': [('Программы', '')]
    })


@login_required
def program_statistics(request, pk):
    """Детальная статистика по программе: пользователи, завершенность, по темам"""
    if not is_subadmin(request.user):
        raise PermissionDenied
    
    program = get_object_or_404(Program, pk=pk)
    
    # Получаем всех оценивателей программы, которые реально оценили хотя бы одну тему
    topic_evals_all = TopicEvaluation.objects.filter(
        program_evaluation__program=program,
        completed_at__isnull=False
    ).select_related('program_evaluation__evaluator', 'topic')
    evaluators = {te.program_evaluation.evaluator for te in topic_evals_all}
    
    # Статистика по каждому оценивателю
    evaluator_stats = []
    for evaluator in evaluators:
        # Находим ProgramEvaluation для этого оценивателя и программы
        pe = ProgramEvaluation.objects.filter(program=program, evaluator=evaluator).first()
        if not pe:
            continue
        
        # Получаем все оценки тем для этого оценивателя (включая ещё не завершённые)
        topic_evals = TopicEvaluation.objects.filter(
            program_evaluation=pe
        ).select_related('topic').prefetch_related('answers__criterion')
        
        total_topics = program.topics.count()
        completed_topics = topic_evals.filter(completed_at__isnull=False).count()
        is_fully_completed = (completed_topics == total_topics and total_topics > 0)
        
        # Детализация по темам
        topics_detail = []
        for topic in program.topics.all().order_by('order', 'id'):
            te = topic_evals.filter(topic=topic).first()
            topics_detail.append({
                'topic': topic,
                'is_completed': te.completed_at is not None if te else False,
                'completed_at': te.completed_at if te and te.completed_at else None,
                'can_re_evaluate': te.can_re_evaluate if te else False,
            })
        
        evaluator_stats.append({
            'evaluator': evaluator,
            'total_topics': total_topics,
            'completed_topics': completed_topics,
            'is_fully_completed': is_fully_completed,
            'topics_detail': topics_detail,
        })
    
    # Общая статистика
    total_evaluators = len(evaluators)
    fully_completed_count = sum(1 for stat in evaluator_stats if stat['is_fully_completed'])
    partially_completed_count = total_evaluators - fully_completed_count
    
    return render(request, 'manager/program_statistics.html', {
        'program': program,
        'evaluator_stats': evaluator_stats,
        'total_evaluators': total_evaluators,
        'fully_completed_count': fully_completed_count,
        'partially_completed_count': partially_completed_count,
        'breadcrumbs': [
            ('Программы', '/manager/'),
            (program.name, f'/manager/program/{program.id}/'),
            ('Статистика', '')
        ]
    })


@login_required
def program_ai_analytics(request, pk):
    if not is_subadmin(request.user):
        raise PermissionDenied

    program = get_object_or_404(Program, pk=pk)
    overview = build_program_overview(program)
    runs = ProgramAIAnalysisRun.objects.filter(program=program).select_related("created_by")
    current_run = runs.first()
    present_run_id = request.GET.get("present_run")
    presentation_steps = [
        "Изучаем методический документ и структуру программы",
        "Сопоставляем оценки закрепленных оценщиков",
        "Проверяем точные, допустимые и условные модальные значения",
        "Собираем аналитический отчет по шаблону",
    ]

    if request.method == "POST":
        form = AIAnalyticsRunForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = form.cleaned_data["methodology_file"]
            file_bytes = uploaded_file.read()
            document_text = extract_document_text(uploaded_file.name, file_bytes)
            context_payload = build_program_ai_context(
                program,
                document_name=uploaded_file.name,
                document_text=document_text,
            )

            run = ProgramAIAnalysisRun(
                program=program,
                created_by=request.user,
                provider_key="stub",
                status=ProgramAIAnalysisRun.STATUS_PENDING,
                methodology_filename=uploaded_file.name,
                methodology_excerpt=document_text[:4000],
                context_json=context_payload,
            )
            run.methodology_file.save(uploaded_file.name, ContentFile(file_bytes), save=False)

            try:
                report_payload = generate_program_ai_report(
                    context_payload,
                    provider_key=run.provider_key,
                )
                context_payload["fake_report"] = report_payload
                run.context_json = context_payload
                run.result_text = report_payload["plain_text"]
                run.status = ProgramAIAnalysisRun.STATUS_COMPLETED
                log_action(
                    request.user,
                    "program_ai_analytics_run",
                    "ProgramAIAnalysisRun",
                    description=f'AI-аналитика для программы "{program.name}"',
                )
            except Exception as exc:
                run.status = ProgramAIAnalysisRun.STATUS_FAILED
                run.error_message = str(exc)

            run.save()
            return redirect(f"{reverse('program_ai_analytics', args=[program.id])}?present_run={run.id}")
    else:
        form = AIAnalyticsRunForm()

    current_report = {}
    should_present_run = False
    if current_run:
        current_report = current_run.context_json.get("fake_report", {})
        should_present_run = str(current_run.id) == str(present_run_id) and current_run.status == ProgramAIAnalysisRun.STATUS_COMPLETED

    return render(request, "manager/program_ai_analytics.html", {
        "program": program,
        "form": form,
        "overview": overview,
        "current_run": current_run,
        "current_report": current_report,
        "should_present_run": should_present_run,
        "presentation_steps": presentation_steps,
        "runs": runs[:10],
        "breadcrumbs": [
            ("Программы", "/manager/"),
            (program.name, f"/manager/program/{program.id}/"),
            ("Аналитика AI", ""),
        ],
    })


@login_required
def download_program_ai_pdf(request, pk):
    if not is_subadmin(request.user):
        raise PermissionDenied

    get_object_or_404(Program, pk=pk)
    pdf_path = Path(settings.BASE_DIR) / "report_from_ai_145_modal.pdf"
    if not pdf_path.exists():
        raise Http404("PDF file not found")

    return FileResponse(
        pdf_path.open("rb"),
        as_attachment=True,
        filename="ai_analytics_demo_report.pdf",
        content_type="application/pdf",
    )


@login_required
@require_POST
def grant_re_evaluation_access(request):
    """Предоставить доступ на переоценку конкретной темы конкретному оценщику"""
    if not is_subadmin(request.user):
        raise PermissionDenied
    
    evaluator_id = request.POST.get('evaluator_id')
    topic_id = request.POST.get('topic_id')
    
    if not evaluator_id or not topic_id:
        return JsonResponse({'success': False, 'error': 'Не указаны обязательные параметры'}, status=400)
    
    try:
        evaluator = EvaluatorSession.objects.get(id=evaluator_id)
        topic = Topic.objects.get(id=topic_id)
        program = topic.program
        
        # Находим ProgramEvaluation
        pe = ProgramEvaluation.objects.filter(evaluator=evaluator, program=program).first()
        if not pe:
            return JsonResponse({'success': False, 'error': 'Оценка программы не найдена'}, status=404)
        
        # Находим или создаем TopicEvaluation
        te, created = TopicEvaluation.objects.get_or_create(
            program_evaluation=pe,
            topic=topic
        )
        
        # Предоставляем доступ на переоценку
        te.can_re_evaluate = True
        te.save()
        
        log_action(
            request.user,
            'grant_re_evaluation',
            'TopicEvaluation',
            te.id,
            description=f'Доступ на переоценку темы "{topic.name}" для {evaluator.full_name}'
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Доступ на переоценку предоставлен {evaluator.full_name} для темы "{topic.name}"'
        })
    except EvaluatorSession.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Оценщик не найден'}, status=404)
    except Topic.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Тема не найдена'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_POST
def grant_re_evaluation_access_program(request):
    """Предоставить доступ на переоценку всех завершенных тем программы для конкретного оценщика"""
    if not is_subadmin(request.user):
        raise PermissionDenied

    evaluator_id = request.POST.get('evaluator_id')
    program_id = request.POST.get('program_id')

    if not evaluator_id or not program_id:
        return JsonResponse({'success': False, 'error': 'Не указаны обязательные параметры'}, status=400)

    try:
        evaluator = EvaluatorSession.objects.get(id=evaluator_id)
        program = Program.objects.get(id=program_id)

        pe = ProgramEvaluation.objects.filter(evaluator=evaluator, program=program).first()
        if not pe:
            return JsonResponse({'success': False, 'error': 'Оценка программы не найдена'}, status=404)

        # Открываем доступ на переоценку по всем уже завершенным темам
        updated_count = TopicEvaluation.objects.filter(
            program_evaluation=pe,
            completed_at__isnull=False
        ).update(can_re_evaluate=True)

        log_action(
            request.user,
            'grant_re_evaluation_program',
            'ProgramEvaluation',
            pe.id,
            description=f'Доступ на переоценку всех завершенных тем программы "{program.name}" для {evaluator.full_name} (обновлено {updated_count} тем)'
        )

        return JsonResponse({
            'success': True,
            'message': f'Доступ на переоценку предоставлен {evaluator.full_name} по всем завершенным темам программы "{program.name}". Обновлено {updated_count} тем.'
        })
    except EvaluatorSession.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Оценщик не найден'}, status=404)
    except Program.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Программа не найдена'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)